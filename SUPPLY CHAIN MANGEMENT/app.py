# app.py — SWAG Product Comparison Dashboard — Version 29.0
# Full production-ready implementation — HTML-matching light theme

import io
import re
import math
import hashlib
import time
import xmlrpc.client
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd
import altair as alt
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

# ─────────────────────────────────────────────────────────────────────────────
# PAGE CONFIG (must be first Streamlit call)
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="SWAG Control Center",
    page_icon="🏷️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 1: SYSTEM KEYS & CANONICAL COLUMN CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────
SYSTEM_KEYS = ["SWAG", "LAROUCHE", "DIFFC", "FASHION_LIMITS"]

PAGE_SIZE = 50

C_SYSTEM       = "System"
C_MODEL        = "Model Code"
C_PRODUCT      = "Product"
C_SALE_PRICE   = "Sale Price"
C_ON_HAND      = "On Hand"
C_BRANCH       = "Branch"
C_LOCATION     = "Location"
C_REFERENCE    = "Reference"
C_TYPE         = "Type"
C_STATE        = "State"
C_FROM         = "From"
C_TO           = "To"
C_QTY          = "Qty"
C_SCHEDULED    = "Scheduled Date"
C_SOLD         = "Sold (30d)"
C_VEL          = "Daily Velocity"
C_DAYS_LEFT    = "Days Left"
C_SUGGEST      = "Suggested Order"
C_PRIORITY     = "Priority"
C_DATE         = "Date"
C_PO           = "PO"
C_SO           = "SO"
C_VENDOR       = "Vendor"
C_CUSTOMER     = "Customer"
C_BRAND_CAT    = "Brand/Category"
C_CATEGORY     = "Category"
C_UNIT_PRICE   = "Unit Price"
C_SUBTOTAL     = "Subtotal"
C_QTY_PURCHASED = "Qty Purchased"

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 2: BILINGUAL SUPPORT
# ─────────────────────────────────────────────────────────────────────────────
def get_lang() -> str:
    return st.session_state.get("lang", "EN")

def t(en: str, ar: str) -> str:
    return ar if get_lang() == "AR" else en

def get_dir() -> str:
    return "rtl" if get_lang() == "AR" else "ltr"

_COL_LABELS = {
    C_SYSTEM:        ("System",           "النظام"),
    C_MODEL:         ("Model Code",       "رمز الموديل"),
    C_PRODUCT:       ("Product",          "المنتج"),
    C_SALE_PRICE:    ("Sale Price",       "سعر البيع"),
    C_ON_HAND:       ("On Hand",          "المتوفر"),
    C_BRANCH:        ("Branch",           "الفرع"),
    C_LOCATION:      ("Location",         "الموقع"),
    C_REFERENCE:     ("Reference",        "المرجع"),
    C_TYPE:          ("Type",             "النوع"),
    C_STATE:         ("State",            "الحالة"),
    C_FROM:          ("From",             "من"),
    C_TO:            ("To",               "إلى"),
    C_QTY:           ("Qty",              "الكمية"),
    C_SCHEDULED:     ("Scheduled Date",   "التاريخ المجدول"),
    C_SOLD:          ("Sold (30d)",       "مباع (30 يوم)"),
    C_VEL:           ("Daily Velocity",   "السرعة اليومية"),
    C_DAYS_LEFT:     ("Days Left",        "الأيام المتبقية"),
    C_SUGGEST:       ("Suggested Order",  "الطلب المقترح"),
    C_PRIORITY:      ("Priority",         "الأولوية"),
    C_DATE:          ("Date",             "التاريخ"),
    C_PO:            ("PO",               "أمر شراء"),
    C_SO:            ("SO",               "أمر بيع"),
    C_VENDOR:        ("Vendor",           "المورد"),
    C_CUSTOMER:      ("Customer",         "العميل"),
    C_BRAND_CAT:     ("Brand/Category",   "العلامة/الفئة"),
    C_CATEGORY:      ("Category",         "الفئة"),
    C_UNIT_PRICE:    ("Unit Price",       "سعر الوحدة"),
    C_SUBTOTAL:      ("Subtotal",         "المجموع الفرعي"),
    C_QTY_PURCHASED: ("Qty Purchased",    "الكمية المشتراة"),
}

def col_label(canonical: str) -> str:
    entry = _COL_LABELS.get(canonical)
    if entry:
        return entry[1] if get_lang() == "AR" else entry[0]
    return canonical

def df_for_display(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    rename = {}
    for canonical, (en, ar) in _COL_LABELS.items():
        if canonical in df.columns:
            rename[canonical] = ar if get_lang() == "AR" else en
    return df.rename(columns=rename) if rename else df.copy()

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 3: SESSION STATE DEFAULTS
# ─────────────────────────────────────────────────────────────────────────────
_DEFAULTS = {
    "authenticated": False,
    "user_email": "",
    "lang": "EN",
    "last_run": None,
    "total_df": None,
    "branch_df": None,
    "transfers_df": None,
    "reorder_df": None,
    "sys_stats": {},
    "search_exact": False,
    "low_stock_thresh": 5,
    "price_history": {},
    "show_transfers": True,
    "show_reorder": True,
    "reorder_mode": "days_cover",
    "reorder_target_days": 30,
    "reorder_max_level": 100,
    "reorder_point": 10,
    "pdf_codes": [],
    "pdf_mode": False,
    "po_analytics_df": None,
    "pc_purch_df": None,
    "pc_stock_df": None,
    "pc_last_code": "",
    "salesanalyticsdf": None,
    "analytics_view": "stock",
    "page_total": 0,
    "page_branch": 0,
    "page_transfers": 0,
    "page_reorder": 0,
    "page_po": 0,
    "page_sales": 0,
}

for _k, _v in _DEFAULTS.items():
    if _k not in st.session_state:
        st.session_state[_k] = _v

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 4: AUTH
# ─────────────────────────────────────────────────────────────────────────────
_TOKEN_SECRET = "swag_v29_prod_2025"

def _make_token(email: str) -> str:
    raw = f"{_TOKEN_SECRET}::{email}"
    return hashlib.sha256(raw.encode()).hexdigest()[:40]

def _verify_token(email: str, token: str) -> bool:
    return bool(email and token and token == _make_token(email))

def restore_session():
    if st.session_state.get("authenticated"):
        return
    try:
        params = st.query_params
        email = params.get("u", "")
        token = params.get("t", "")
        if email and token and _verify_token(email, token):
            st.session_state.authenticated = True
            st.session_state.user_email = email
    except Exception:
        pass

def do_logout():
    for key in list(st.session_state.keys()):
        del st.session_state[key]
    try:
        st.query_params.clear()
    except Exception:
        pass
    st.rerun()

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 5: ODOO XML-RPC HELPERS
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_resource
def _proxy(url: str, ep: str):
    return xmlrpc.client.ServerProxy(f"{url.rstrip('/')}/xmlrpc/2/{ep}", allow_none=True)

@st.cache_data(ttl=3600, show_spinner=False)
def _auth(url: str, db: str, user: str, key: str):
    try:
        uid = _proxy(url, "common").authenticate(db, user, key, {})
        return uid if (uid and isinstance(uid, int) and uid > 0) else None
    except Exception:
        return None

def _x(url, db, uid, key, model, method, domain, kw=None):
    if kw is None:
        kw = {}
    return _proxy(url, "object").execute_kw(db, uid, key, model, method, domain, kw)

def _domain(codes: list, exact: bool) -> list:
    if not codes:
        return []
    if exact:
        return [("default_code", "in", codes)]
    clauses = [("default_code", "=ilike", f"{c}%") for c in codes]
    if len(clauses) == 1:
        return [clauses[0]]
    domain = []
    for _ in range(len(clauses) - 1):
        domain.append("|")
    domain.extend(clauses)
    return domain

def _get_conn(key: str):
    cfg = st.secrets.get(key, {})
    url = cfg.get("url", "").rstrip("/")
    db = cfg.get("db", "")
    user = cfg.get("user", "")
    api_key = cfg.get("api_key", "")
    if not url or not db or not user or not api_key:
        return None, None, None, None, key, f"[{key}] Missing config."
    uid = _auth(url, db, user, api_key)
    if not uid:
        return url, db, None, api_key, key, f"[{key}] Auth failed."
    name_en = cfg.get("name", key)
    name_ar = cfg.get("name_ar", name_en)
    display_name = name_ar if get_lang() == "AR" else name_en
    return url, db, uid, api_key, display_name, None

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 6: NUMERIC HELPER
# ─────────────────────────────────────────────────────────────────────────────
def _to_num(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce").fillna(0)

def _safe_val(val):
    try:
        return float(val)
    except Exception:
        return 0.0

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 7: PDF INVOICE PARSING
# ─────────────────────────────────────────────────────────────────────────────
_EXCLUDE = {
    "SR", "VAT", "TAX", "PCS", "QTY", "NO", "REF", "INV", "PO", "SO",
    "DO", "ID", "EN", "AR", "PDF", "AED", "SAR", "USD", "KWD", "OMR",
    "BHD", "JOD", "EGP", "TRY",
}

_PDF_PATTERNS = [
    re.compile(r"\[([A-Z0-9][A-Z0-9\-]{2,})\]"),
    re.compile(r"^SR\s+([A-Z0-9][A-Z0-9\-]{2,})", re.MULTILINE),
    re.compile(r"\b([A-Z]{2,}[0-9]{2,}[A-Z0-9\-]*)\b"),
]

def parse_invoice_pdf_cached(file_bytes: bytes) -> list:
    try:
        import pypdf
        reader = pypdf.PdfReader(io.BytesIO(file_bytes))
    except Exception:
        return []
    text = ""
    for page in reader.pages:
        try:
            text += (page.extract_text() or "") + "\n"
        except Exception:
            pass
    found = []
    seen = set()
    seq = 0
    for pat in _PDF_PATTERNS:
        for m in pat.finditer(text):
            code = m.group(1).strip().upper()
            if code in _EXCLUDE or len(code) < 3:
                continue
            if code not in seen:
                seen.add(code)
                found.append({"sequence": seq, "code": code})
                seq += 1
    return found

def extract_base_model(code: str) -> str:
    cleaned = re.sub(r"[-_](XS|S|M|L|XL|XXL|2XL|3XL|XXXL|\d{2,3})$", "", code.upper().strip())
    return cleaned

def get_unique_base_models(raw: list) -> list:
    seen = set()
    result = []
    for item in raw:
        base = extract_base_model(item["code"])
        if base not in seen:
            seen.add(base)
            result.append(base)
    return result

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 8: DATA FETCHING
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_data(ttl=1800, show_spinner=False)
def _fetch_stock_one(key, codes_tuple, exact, low_thresh, show_transfers, show_reorder,
                     reorder_mode, reorder_target_days, reorder_max_level, reorder_point):
    url, db, uid, ak, name, err = _get_conn(key)
    if err:
        return [], [], [], [], {"system": name, "level": "error", "msg": err}
    codes = list(codes_tuple)
    try:
        prod_domain = _domain(codes, exact) if codes else []
        templates = _x(url, db, uid, ak, "product.template", "search_read",
                        [prod_domain if prod_domain else []],
                        {"fields": ["id", "name", "default_code", "list_price", "categ_id"], "limit": 5000})
        if not templates:
            return [], [], [], [], {"system": name, "level": "ok", "msg": "No products found."}
        tmpl_map = {t["id"]: t for t in templates}
        tmpl_ids = list(tmpl_map.keys())
        variants = _x(url, db, uid, ak, "product.product", "search_read",
                       [[("product_tmpl_id", "in", tmpl_ids)]],
                       {"fields": ["id", "product_tmpl_id"], "limit": 50000})
        var_to_tmpl = {}
        for v in variants:
            raw = v.get("product_tmpl_id")
            tid = raw[0] if isinstance(raw, list) else raw
            var_to_tmpl[v["id"]] = tid
        var_ids = list(var_to_tmpl.keys())
        tmpl_qty = {}
        branch_rows = []
        if var_ids:
            quants = _x(url, db, uid, ak, "stock.quant", "search_read",
                         [[("product_id", "in", var_ids), ("location_id.usage", "=", "internal")]],
                         {"fields": ["product_id", "location_id", "quantity"], "limit": 50000})
            for q in quants:
                pid_raw = q.get("product_id")
                vid = pid_raw[0] if isinstance(pid_raw, list) else pid_raw
                tid = var_to_tmpl.get(vid)
                if tid is None:
                    continue
                qty = float(q.get("quantity") or 0)
                tmpl_qty[tid] = tmpl_qty.get(tid, 0) + qty
                loc_raw = q.get("location_id")
                loc_name = loc_raw[1] if isinstance(loc_raw, list) and len(loc_raw) > 1 else str(loc_raw or "")
                mc = (tmpl_map.get(tid, {}).get("default_code") or "").strip()
                if mc:
                    branch_rows.append({C_SYSTEM: name, C_BRANCH: loc_name, C_MODEL: mc, C_ON_HAND: qty})
        date_30_ago = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
        vel_map = {}
        if var_ids:
            try:
                so_lines = _x(url, db, uid, ak, "sale.order.line", "search_read",
                               [[("product_id", "in", var_ids),
                                 ("order_id.date_order", ">=", f"{date_30_ago} 00:00:00"),
                                 ("order_id.state", "in", ["sale", "done"])]],
                               {"fields": ["product_id", "product_uom_qty"], "limit": 50000})
                for sl in so_lines:
                    pid_raw = sl.get("product_id")
                    vid = pid_raw[0] if isinstance(pid_raw, list) else pid_raw
                    tid = var_to_tmpl.get(vid)
                    if tid:
                        vel_map[tid] = vel_map.get(tid, 0) + float(sl.get("product_uom_qty") or 0)
            except Exception:
                pass
        total_rows = []
        for tid, tmpl in tmpl_map.items():
            mc = (tmpl.get("default_code") or "").strip()
            categ_raw = tmpl.get("categ_id")
            category = categ_raw[1] if isinstance(categ_raw, list) and len(categ_raw) > 1 else ""
            on_hand = tmpl_qty.get(tid, 0)
            sale_price = float(tmpl.get("list_price") or 0)
            sold_30 = vel_map.get(tid, 0)
            daily_vel = sold_30 / 30.0
            total_rows.append({
                C_SYSTEM: name, C_MODEL: mc, C_PRODUCT: tmpl.get("name", ""),
                C_SALE_PRICE: sale_price, C_ON_HAND: on_hand,
                C_SOLD: sold_30, C_VEL: round(daily_vel, 3), C_CATEGORY: category,
            })
        reorder_rows = []
        if show_reorder:
            for tid, tmpl in tmpl_map.items():
                mc = (tmpl.get("default_code") or "").strip()
                on_hand = tmpl_qty.get(tid, 0)
                sold_30 = vel_map.get(tid, 0)
                daily_vel = sold_30 / 30.0
                if reorder_mode == "days_cover":
                    days_left = (on_hand / daily_vel) if daily_vel > 0 else 999
                    suggest = max(0, round(daily_vel * reorder_target_days - on_hand))
                    priority = ("🔴 Critical" if days_left < 7 else "🟡 Low" if days_left < 14 else "🟢 OK")
                else:
                    days_left = (on_hand / daily_vel) if daily_vel > 0 else 999
                    suggest = max(0, reorder_max_level - on_hand) if on_hand <= reorder_point else 0
                    priority = "🔴 Critical" if on_hand <= reorder_point else "🟢 OK"
                if suggest > 0 or on_hand <= low_thresh:
                    reorder_rows.append({
                        C_SYSTEM: name, C_MODEL: mc, C_PRODUCT: tmpl.get("name", ""),
                        C_ON_HAND: on_hand, C_SOLD: sold_30, C_VEL: round(daily_vel, 3),
                        C_DAYS_LEFT: round(days_left, 1) if days_left < 999 else "∞",
                        C_SUGGEST: suggest, C_PRIORITY: priority,
                    })
        transfer_rows = []
        if show_transfers and var_ids:
            try:
                moves = _x(url, db, uid, ak, "stock.move", "search_read",
                            [[("product_id", "in", var_ids),
                              ("state", "not in", ["cancel", "done"]),
                              ("picking_id.picking_type_code", "=", "internal")]],
                            {"fields": ["product_id", "product_uom_qty", "state",
                                        "location_id", "location_dest_id", "reference",
                                        "date", "picking_id"], "limit": 5000})
                for mv in moves:
                    pid_raw = mv.get("product_id")
                    vid = pid_raw[0] if isinstance(pid_raw, list) else pid_raw
                    tid = var_to_tmpl.get(vid)
                    mc = (tmpl_map.get(tid, {}).get("default_code") or "").strip() if tid else ""
                    loc_from = mv.get("location_id")
                    loc_to = mv.get("location_dest_id")
                    transfer_rows.append({
                        C_SYSTEM: name, C_MODEL: mc,
                        C_PRODUCT: tmpl_map.get(tid, {}).get("name", "") if tid else "",
                        C_QTY: float(mv.get("product_uom_qty") or 0),
                        C_STATE: mv.get("state", ""),
                        C_FROM: loc_from[1] if isinstance(loc_from, list) and len(loc_from) > 1 else "",
                        C_TO: loc_to[1] if isinstance(loc_to, list) and len(loc_to) > 1 else "",
                        C_REFERENCE: mv.get("reference", ""),
                        C_SCHEDULED: str(mv.get("date", ""))[:10],
                    })
            except Exception:
                pass
        return total_rows, branch_rows, transfer_rows, reorder_rows, {"system": name, "level": "ok", "msg": f"Loaded {len(total_rows)} products."}
    except Exception as e:
        return [], [], [], [], {"system": name, "level": "error", "msg": f"{type(e).__name__}: {e}"}


def fetch_all_data(codes, exact, low_stock_thresh, show_transfers, show_reorder,
                   reorder_mode, reorder_target_days, reorder_max_level, reorder_point):
    codes_tuple = tuple(sorted(set(codes))) if codes else ()
    all_total, all_branch, all_transfers, all_reorder, sys_stats = [], [], [], [], {}
    with ThreadPoolExecutor(max_workers=4) as ex:
        futs = {ex.submit(_fetch_stock_one, k, codes_tuple, exact, low_stock_thresh,
                          show_transfers, show_reorder, reorder_mode,
                          reorder_target_days, reorder_max_level, reorder_point): k for k in SYSTEM_KEYS}
        for f in as_completed(futs):
            key = futs[f]
            tr, br, tf, ro, stat = f.result()
            all_total.extend(tr); all_branch.extend(br)
            all_transfers.extend(tf); all_reorder.extend(ro)
            sys_stats[key] = stat
    num_cols_total = [C_SALE_PRICE, C_ON_HAND, C_SOLD, C_VEL]
    total_df = pd.DataFrame(all_total) if all_total else pd.DataFrame(
        columns=[C_SYSTEM, C_MODEL, C_PRODUCT, C_SALE_PRICE, C_ON_HAND, C_SOLD, C_VEL, C_CATEGORY])
    for c in num_cols_total:
        if c in total_df.columns:
            total_df[c] = _to_num(total_df[c])
    branch_df = pd.DataFrame(all_branch) if all_branch else pd.DataFrame(
        columns=[C_SYSTEM, C_BRANCH, C_MODEL, C_ON_HAND])
    if C_ON_HAND in branch_df.columns:
        branch_df[C_ON_HAND] = _to_num(branch_df[C_ON_HAND])
    transfers_df = pd.DataFrame(all_transfers) if all_transfers else pd.DataFrame(
        columns=[C_SYSTEM, C_MODEL, C_PRODUCT, C_QTY, C_STATE, C_FROM, C_TO, C_REFERENCE, C_SCHEDULED])
    if C_QTY in transfers_df.columns:
        transfers_df[C_QTY] = _to_num(transfers_df[C_QTY])
    reorder_df = pd.DataFrame(all_reorder) if all_reorder else pd.DataFrame(
        columns=[C_SYSTEM, C_MODEL, C_PRODUCT, C_ON_HAND, C_SOLD, C_VEL, C_DAYS_LEFT, C_SUGGEST, C_PRIORITY])
    return total_df, branch_df, transfers_df, reorder_df, sys_stats

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 9: PURCHASE HISTORY FETCHING
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_data(ttl=1800, show_spinner=False)
def fetch_purchase_history_for_system(system_key, model_codes, date_from, date_to):
    url, db, uid, ak, name, err = _get_conn(system_key)
    _empty = pd.DataFrame(columns=[C_SYSTEM, C_DATE, C_PO, C_VENDOR, C_PRODUCT, C_MODEL,
                                    C_CATEGORY, C_BRAND_CAT, C_QTY_PURCHASED, C_UNIT_PRICE, C_SUBTOTAL, C_STATE])
    if err:
        return _empty
    try:
        po_domain = [("date_approve", ">=", f"{date_from} 00:00:00"),
                     ("date_approve", "<=", f"{date_to} 23:59:59"),
                     ("state", "in", ["purchase", "done"])]
        pos_list = _x(url, db, uid, ak, "purchase.order", "search_read",
                       [po_domain], {"fields": ["id", "name", "partner_id", "date_approve", "state"], "limit": 2000})
        if not pos_list:
            return _empty
        po_ids = [p["id"] for p in pos_list]
        po_map = {p["id"]: p for p in pos_list}
        line_domain = [("order_id", "in", po_ids)]
        if model_codes:
            line_domain.append(("product_id.default_code", "in", list(model_codes)))
        lines = _x(url, db, uid, ak, "purchase.order.line", "search_read",
                    [line_domain],
                    {"fields": ["order_id", "product_id", "product_qty", "price_unit", "price_subtotal"], "limit": 20000})
        if not lines:
            return _empty
        prod_ids = list({l["product_id"][0] for l in lines if isinstance(l.get("product_id"), list)})
        products = _x(url, db, uid, ak, "product.product", "search_read",
                       [[("id", "in", prod_ids)]],
                       {"fields": ["id", "default_code", "name", "categ_id"], "limit": len(prod_ids) + 10})
        prod_map = {p["id"]: p for p in products}
        rows = []
        for line in lines:
            oid_raw = line.get("order_id"); oid = oid_raw[0] if isinstance(oid_raw, list) else oid_raw
            po = po_map.get(oid, {})
            pid_raw = line.get("product_id"); pid = pid_raw[0] if isinstance(pid_raw, list) else pid_raw
            prod = prod_map.get(pid, {})
            mc = (prod.get("default_code") or "").strip()
            categ_raw = prod.get("categ_id")
            category = categ_raw[1] if isinstance(categ_raw, list) and len(categ_raw) > 1 else ""
            partner_raw = po.get("partner_id")
            vendor = partner_raw[1] if isinstance(partner_raw, list) and len(partner_raw) > 1 else ""
            rows.append({C_SYSTEM: name, C_DATE: str(po.get("date_approve", ""))[:10],
                         C_PO: po.get("name", ""), C_VENDOR: vendor,
                         C_PRODUCT: prod.get("name", ""), C_MODEL: mc,
                         C_CATEGORY: category, C_BRAND_CAT: category,
                         C_QTY_PURCHASED: float(line.get("product_qty") or 0),
                         C_UNIT_PRICE: float(line.get("price_unit") or 0),
                         C_SUBTOTAL: float(line.get("price_subtotal") or 0),
                         C_STATE: po.get("state", "")})
        if not rows:
            return _empty
        df = pd.DataFrame(rows)
        df[C_DATE] = pd.to_datetime(df[C_DATE], errors="coerce")
        for c in [C_QTY_PURCHASED, C_UNIT_PRICE, C_SUBTOTAL]:
            df[c] = _to_num(df[c])
        return df.sort_values(C_DATE, ascending=False).reset_index(drop=True)
    except Exception:
        return _empty

def fetch_all_systems_purchase_history(model_codes, date_from, date_to):
    codes_tuple = tuple(sorted(set(model_codes))) if model_codes else ()
    results = []
    with ThreadPoolExecutor(max_workers=4) as ex:
        futs = [ex.submit(fetch_purchase_history_for_system, k, codes_tuple, date_from, date_to) for k in SYSTEM_KEYS]
        for f in as_completed(futs):
            df = f.result()
            if df is not None and not df.empty:
                results.append(df)
    if not results:
        return pd.DataFrame()
    combined = pd.concat(results, ignore_index=True)
    combined[C_DATE] = pd.to_datetime(combined[C_DATE], errors="coerce")
    return combined.sort_values(C_DATE, ascending=False).reset_index(drop=True)

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 10: SALES HISTORY FETCHING
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_data(ttl=1800, show_spinner=False)
def fetch_sales_history_for_system(system_key, model_codes, date_from, date_to):
    url, db, uid, ak, name, err = _get_conn(system_key)
    _empty = pd.DataFrame(columns=[C_SYSTEM, C_DATE, C_SO, C_CUSTOMER, C_PRODUCT, C_MODEL,
                                    C_CATEGORY, C_BRAND_CAT, C_QTY, C_UNIT_PRICE, C_SUBTOTAL, C_STATE])
    if err:
        return _empty
    try:
        so_domain = [("date_order", ">=", f"{date_from} 00:00:00"),
                     ("date_order", "<=", f"{date_to} 23:59:59"),
                     ("state", "in", ["sale", "done"])]
        orders = _x(url, db, uid, ak, "sale.order", "search_read",
                     [so_domain],
                     {"fields": ["id", "name", "date_order", "partner_id", "state", "order_line"], "limit": 5000})
        if not orders:
            return _empty
        order_map = {o["id"]: o for o in orders}
        line_ids = []
        for o in orders:
            if o.get("order_line"):
                line_ids.extend(o["order_line"])
        if not line_ids:
            return _empty
        line_domain = [("id", "in", line_ids)]
        if model_codes:
            line_domain.append(("product_id.default_code", "in", list(model_codes)))
        lines = _x(url, db, uid, ak, "sale.order.line", "search_read",
                    [line_domain],
                    {"fields": ["order_id", "product_id", "product_uom_qty", "price_unit", "price_subtotal"], "limit": 20000})
        if not lines:
            return _empty
        prod_ids = list({l["product_id"][0] for l in lines if isinstance(l.get("product_id"), list)})
        products = _x(url, db, uid, ak, "product.product", "search_read",
                       [[("id", "in", prod_ids)]],
                       {"fields": ["id", "default_code", "name", "categ_id"], "limit": len(prod_ids) + 10})
        prod_map = {p["id"]: p for p in products}
        rows = []
        for line in lines:
            oid_raw = line.get("order_id"); oid = oid_raw[0] if isinstance(oid_raw, list) else oid_raw
            order = order_map.get(oid, {})
            pid_raw = line.get("product_id"); pid = pid_raw[0] if isinstance(pid_raw, list) else pid_raw
            prod = prod_map.get(pid, {})
            mc = (prod.get("default_code") or "").strip()
            categ_raw = prod.get("categ_id")
            category = categ_raw[1] if isinstance(categ_raw, list) and len(categ_raw) > 1 else ""
            partner_raw = order.get("partner_id")
            customer = partner_raw[1] if isinstance(partner_raw, list) and len(partner_raw) > 1 else ""
            rows.append({C_SYSTEM: name, C_DATE: str(order.get("date_order", ""))[:10],
                         C_SO: order.get("name", ""), C_CUSTOMER: customer,
                         C_PRODUCT: prod.get("name", ""), C_MODEL: mc,
                         C_CATEGORY: category, C_BRAND_CAT: category,
                         C_QTY: float(line.get("product_uom_qty") or 0),
                         C_UNIT_PRICE: float(line.get("price_unit") or 0),
                         C_SUBTOTAL: float(line.get("price_subtotal") or 0),
                         C_STATE: order.get("state", "")})
        if not rows:
            return _empty
        df = pd.DataFrame(rows)
        df[C_DATE] = pd.to_datetime(df[C_DATE], errors="coerce")
        for c in [C_QTY, C_UNIT_PRICE, C_SUBTOTAL]:
            df[c] = _to_num(df[c])
        return df.sort_values(C_DATE, ascending=False).reset_index(drop=True)
    except Exception:
        return _empty

def fetch_all_systems_sales_history(model_codes, date_from, date_to):
    codes_tuple = tuple(sorted(set(model_codes))) if model_codes else ()
    results = []
    with ThreadPoolExecutor(max_workers=4) as ex:
        futs = [ex.submit(fetch_sales_history_for_system, k, codes_tuple, date_from, date_to) for k in SYSTEM_KEYS]
        for f in as_completed(futs):
            df = f.result()
            if df is not None and not df.empty:
                results.append(df)
    if not results:
        return pd.DataFrame()
    combined = pd.concat(results, ignore_index=True)
    combined[C_DATE] = pd.to_datetime(combined[C_DATE], errors="coerce")
    return combined.sort_values(C_DATE, ascending=False).reset_index(drop=True)

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 11: EXCEL / CSV EXPORT
# ─────────────────────────────────────────────────────────────────────────────
def _style_worksheet(ws, df_clean, lang="EN"):
    try:
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.formatting.rule import DataBarRule, ColorScaleRule, CellIsRule
        header_fill = PatternFill("solid", fgColor="0EA5E9")
        alt_fill    = PatternFill("solid", fgColor="F0F9FF")
        zero_fill   = PatternFill("solid", fgColor="FEE2E2")
        total_fill  = PatternFill("solid", fgColor="0369A1")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        total_font  = Font(bold=True, color="FFFFFF", size=10)
        thin = Border(
            left=Side(style="thin", color="E5E7EB"),
            right=Side(style="thin", color="E5E7EB"),
            top=Side(style="thin", color="E5E7EB"),
            bottom=Side(style="thin", color="E5E7EB"),
        )
        if lang == "AR":
            ws.sheet_view.rightToLeft = True
        for col_idx, col_name in enumerate(df_clean.columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = col_name
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(12, min(35, len(str(col_name)) + 4))
        total_cols = list(df_clean.columns)
        on_hand_col_idx = (total_cols.index(C_ON_HAND) + 1) if C_ON_HAND in total_cols else None
        totals = {}
        for row_idx, (_, row) in enumerate(df_clean.iterrows(), 2):
            on_hand_val = 0
            for col_idx, val in enumerate(row.values, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = "" if isinstance(val, float) and val != val else val
                if col_idx == on_hand_col_idx:
                    on_hand_val = float(val) if val and val == val else 0
                if row_idx % 2 == 0:
                    cell.fill = alt_fill
                cell.border = thin
                cell.font = Font(color="111827", size=9)
                cell.alignment = Alignment(horizontal="right" if lang == "AR" else "left", vertical="center")
                try:
                    totals[col_idx] = totals.get(col_idx, 0) + float(val)
                except (TypeError, ValueError):
                    pass
            if on_hand_val == 0:
                for col_idx2 in range(1, len(total_cols) + 1):
                    ws.cell(row=row_idx, column=col_idx2).fill = zero_fill
        total_row = ws.max_row + 1
        ws.cell(row=total_row, column=1).value = "TOTAL"
        for col_idx in range(1, len(total_cols) + 1):
            cell = ws.cell(row=total_row, column=col_idx)
            cell.fill = total_fill
            cell.font = total_font
            cell.border = thin
            if col_idx in totals:
                cell.value = round(totals[col_idx], 2)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
    except Exception:
        pass

def to_csv(df): return df.to_csv(index=False).encode("utf-8-sig")

def to_excel(df):
    if df is None or df.empty:
        out = io.BytesIO()
        with pd.ExcelWriter(out, engine="openpyxl") as w:
            pd.DataFrame({"Message": ["No data"]}).to_excel(w, sheet_name="Data", index=False)
        out.seek(0); return out.getvalue()
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as w:
        df.to_excel(w, sheet_name="Data", index=False)
        _style_worksheet(w.sheets["Data"], df, get_lang())
    out.seek(0); return out.getvalue()

def dl_name(prefix, ext):
    return f"{prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{ext}"

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 12: PAGINATION
# ─────────────────────────────────────────────────────────────────────────────
def paginate_df(df, page_key, page_size=PAGE_SIZE):
    if df is None or df.empty:
        return df, 1, 0
    total = len(df)
    total_pages = max(1, math.ceil(total / page_size))
    current = min(st.session_state.get(page_key, 0), total_pages - 1)
    st.session_state[page_key] = current
    start = current * page_size
    end = min(start + page_size, total)
    page_df = df.iloc[start:end].copy()

    info_html = (
        f"<div style='text-align:center;margin:6px 0;font-size:0.78rem;color:#6b7280;'>"
        f"{t('Showing','عرض')} {start+1}–{end} {t('of','من')} {total} &nbsp;|&nbsp; "
        f"<span style='background:#f0f9ff;border:1px solid #bae6fd;padding:2px 10px;border-radius:12px;color:#0369a1;font-weight:600;'>"
        f"{t('Page','صفحة')} {current+1}/{total_pages}</span></div>"
    )
    st.markdown(info_html, unsafe_allow_html=True)
    c1, c2, _, c3, c4 = st.columns([1, 1, 3, 1, 1])
    if c1.button("⏮", key=f"{page_key}_first", use_container_width=True):
        st.session_state[page_key] = 0; st.rerun()
    if c2.button("◀", key=f"{page_key}_prev", use_container_width=True):
        st.session_state[page_key] = max(0, current - 1); st.rerun()
    if c3.button("▶", key=f"{page_key}_next", use_container_width=True):
        st.session_state[page_key] = min(total_pages - 1, current + 1); st.rerun()
    if c4.button("⏭", key=f"{page_key}_last", use_container_width=True):
        st.session_state[page_key] = total_pages - 1; st.rerun()
    return page_df, total_pages, current

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 13: TABLE RENDERING  ── Light theme matching HTML
# ─────────────────────────────────────────────────────────────────────────────
def _days_left_pill(val):
    """Return HTML pill for days-left value."""
    try:
        if str(val) == "∞" or val is None:
            return f"<span class='pill-soft'>∞ days</span>"
        d = float(val)
        if d < 7:
            return f"<span class='pill-danger'>🔴 {d:.0f}d left</span>"
        if d < 14:
            return f"<span class='pill-warning'>🟡 {d:.0f}d left</span>"
        return f"<span class='pill-ok'>🟢 {d:.0f}d left</span>"
    except Exception:
        return f"<span class='pill-soft'>{val}</span>"

def _priority_pill(val):
    v = str(val)
    if "Critical" in v:
        return f"<span class='pill-danger'>{v}</span>"
    if "Low" in v:
        return f"<span class='pill-warning'>{v}</span>"
    return f"<span class='pill-ok'>{v}</span>"

def _status_pill(qty):
    if float(qty) == 0:
        return "<span class='pill-danger'>Out of stock</span>"
    return "<span class='pill-ok'>Healthy</span>"

def render_premium_table(df: pd.DataFrame, thresh: int = 0, accent_cols: list = None):
    if df is None or df.empty:
        st.markdown(
            f"<div class='empty-state'>ℹ️ {t('No data to display.','لا توجد بيانات للعرض.')}</div>",
            unsafe_allow_html=True,
        )
        return
    if accent_cols is None:
        accent_cols = []
    is_rtl = get_lang() == "AR"
    dir_attr = 'dir="rtl"' if is_rtl else ''
    text_align = "right" if is_rtl else "left"

    # Detect special display columns
    days_left_col = col_label(C_DAYS_LEFT)
    priority_col = col_label(C_PRIORITY)
    on_hand_col = col_label(C_ON_HAND)

    header_html = "".join(
        f"<th>{c}</th>" for c in df.columns
    )

    rows_html = ""
    for i, (_, row) in enumerate(df.iterrows()):
        # Check on-hand value for row highlight
        try:
            oh_val = float(row.get(on_hand_col, row.get(C_ON_HAND, 1)))
        except Exception:
            oh_val = 1
        row_class = "row-zero" if oh_val == 0 else ("row-even" if i % 2 == 0 else "row-odd")

        cells = ""
        for col_name, val in zip(df.columns, row.values):
            # Special pill rendering
            if col_name == days_left_col or col_name == C_DAYS_LEFT:
                cell_content = _days_left_pill(val)
            elif col_name == priority_col or col_name == C_PRIORITY:
                cell_content = _priority_pill(val)
            elif col_name in accent_cols:
                cell_content = f"<strong style='color:#0ea5e9;'>{val}</strong>"
            else:
                cell_content = str(val) if val is not None else ""
            cells += f"<td style='text-align:{text_align};'>{cell_content}</td>"
        rows_html += f"<tr class='{row_class}'>{cells}</tr>"

    table_html = (
        f"<div class='table-wrap' {dir_attr}>"
        f"<table><thead><tr>{header_html}</tr></thead>"
        f"<tbody>{rows_html}</tbody></table></div>"
    )
    st.markdown(table_html, unsafe_allow_html=True)

def display_df(df, thresh=0, table_key=None):
    if df is None or df.empty:
        render_premium_table(df, thresh); return
    key = table_key or f"tbl_{abs(hash(str(df.columns.tolist()))) % 10**8}"
    page_df, _, _ = paginate_df(df, key)
    display = df_for_display(page_df)
    render_premium_table(display, thresh)

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 14: UI COMPONENT HELPERS  ── Light theme matching HTML
# ─────────────────────────────────────────────────────────────────────────────
def _kpi_card(icon, value, label, meta="", color="#0ea5e9"):
    return (
        f"<div class='kpi-card'>"
        f"<div class='kpi-icon'>{icon}</div>"
        f"<div class='kpi-value' style='color:{color};'>{value}</div>"
        f"<div class='kpi-label'>{label}</div>"
        f"<div class='kpi-meta'>{meta}</div>"
        f"</div>"
    )

def _render_kpi_grid(cards):
    n = len(cards)
    html = f"<div class='kpi-row' style='grid-template-columns:repeat({min(n,4)},1fr);'>"
    html += "".join(cards)
    html += "</div>"
    st.markdown(html, unsafe_allow_html=True)

def _section_header(title, icon, subtitle=""):
    sub = f"<div class='section-sub'>{subtitle}</div>" if subtitle else ""
    st.markdown(
        f"<div class='section-header'>"
        f"<span class='section-icon'>{icon}</span>"
        f"<div><div class='section-title'>{title}</div>{sub}</div>"
        f"</div>",
        unsafe_allow_html=True,
    )

def _mini_card(label, value, meta="", value_color="#111827"):
    return (
        f"<div class='mini-card'>"
        f"<div class='mini-label'>{label}</div>"
        f"<div class='mini-value' style='color:{value_color};'>{value}</div>"
        f"<div class='mini-meta'>{meta}</div>"
        f"</div>"
    )

def _render_mini_grid(cards):
    html = "<div class='mini-grid'>" + "".join(cards) + "</div>"
    st.markdown(html, unsafe_allow_html=True)

def _divider():
    st.markdown("<hr class='swag-divider'>", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 15: CHART HELPERS (Plotly light theme)
# ─────────────────────────────────────────────────────────────────────────────
_COLORS = ["#3b82f6","#22c55e","#f97316","#a855f7","#0ea5e9","#f59e0b","#ef4444","#10b981"]

def _apply_light_theme(fig):
    fig.update_layout(
        paper_bgcolor="rgba(255,255,255,0)",
        plot_bgcolor="rgba(255,255,255,0)",
        font=dict(color="#374151", family="'DM Sans', sans-serif", size=11),
        margin=dict(l=20, r=20, t=40, b=20),
        legend=dict(bgcolor="rgba(255,255,255,0.9)", bordercolor="#e5e7eb", borderwidth=1),
        title_font=dict(size=13, color="#111827"),
    )
    fig.update_xaxes(gridcolor="#f3f4f6", linecolor="#e5e7eb", tickfont=dict(size=10, color="#6b7280"))
    fig.update_yaxes(gridcolor="#f3f4f6", linecolor="#e5e7eb", tickfont=dict(size=10, color="#9ca3af"))
    return fig

def _top10_bar(df, col_name, val_col, title, color="#3b82f6"):
    if df is None or df.empty or col_name not in df.columns or val_col not in df.columns:
        return None
    agg = df.groupby(col_name)[val_col].sum().reset_index().nlargest(10, val_col)
    agg[col_name] = agg[col_name].astype(str).str[:28]
    fig = px.bar(agg, x=val_col, y=col_name, orientation="h",
                 title=title, color_discrete_sequence=[color])
    fig.update_traces(marker_line_width=0, marker_cornerradius=4)
    return _apply_light_theme(fig)

def _donut_chart(df, names_col, values_col, title, colors=None):
    if df is None or df.empty:
        return None
    if colors is None:
        colors = _COLORS
    fig = px.pie(df, names=names_col, values=values_col, hole=0.58,
                 title=title, color_discrete_sequence=colors)
    fig.update_traces(textposition="inside", textinfo="percent+label",
                      marker=dict(line=dict(color="#ffffff", width=2)))
    return _apply_light_theme(fig)

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 16: GLOBAL CSS  ── Matches HTML design exactly
# ─────────────────────────────────────────────────────────────────────────────
def _build_css() -> str:
    is_rtl = get_lang() == "AR"
    body_dir = "rtl" if is_rtl else "ltr"
    return f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700&family=DM+Mono:wght@400;500&display=swap');

/* ── Reset & base ── */
*, *::before, *::after {{ box-sizing: border-box; }}
html, body {{ direction: {body_dir}; }}

/* ── App background ── */
.stApp {{
    background:
        radial-gradient(circle at 0% 0%, rgba(59,130,246,0.10) 0, transparent 45%),
        radial-gradient(circle at 100% 100%, rgba(34,197,94,0.10) 0, transparent 45%),
        #f3f5fb !important;
    color: #111827;
    font-family: 'DM Sans', system-ui, sans-serif !important;
}}

/* ── Sidebar ── */
section[data-testid="stSidebar"] {{
    background: #ffffff !important;
    border-right: 1px solid #e5e7eb !important;
}}
section[data-testid="stSidebar"] * {{
    color: #374151 !important;
    font-family: 'DM Sans', sans-serif !important;
}}
section[data-testid="stSidebar"] .stSelectbox > div > div {{
    background: #f9fafb !important;
    border-color: #e5e7eb !important;
    color: #111827 !important;
}}

/* ── Inputs ── */
.stTextInput input, .stNumberInput input, .stTextArea textarea {{
    background: #f9fafb !important;
    border: 1px solid #e5e7eb !important;
    border-radius: 10px !important;
    color: #111827 !important;
    font-size: 0.875rem !important;
    font-family: 'DM Sans', sans-serif !important;
}}
.stTextInput input:focus, .stTextArea textarea:focus {{
    border-color: #0ea5e9 !important;
    box-shadow: 0 0 0 3px rgba(14,165,233,0.12) !important;
}}
label {{ color: #6b7280 !important; font-size: 0.78rem !important; font-weight: 500 !important; }}

/* ── Buttons ── */
.stButton > button {{
    border-radius: 10px !important;
    border: 1px solid #e5e7eb !important;
    background: #ffffff !important;
    color: #374151 !important;
    font-weight: 500 !important;
    font-family: 'DM Sans', sans-serif !important;
    font-size: 0.82rem !important;
    transition: all 0.16s ease !important;
    box-shadow: 0 1px 3px rgba(0,0,0,0.06) !important;
}}
.stButton > button:hover {{
    background: #f9fafb !important;
    border-color: #0ea5e9 !important;
    color: #0ea5e9 !important;
    transform: translateY(-1px);
    box-shadow: 0 4px 12px rgba(14,165,233,0.15) !important;
}}
.stButton > button[kind="primary"] {{
    background: linear-gradient(135deg, #0ea5e9 0%, #2563eb 100%) !important;
    color: #fff !important;
    border: none !important;
    box-shadow: 0 4px 14px rgba(37,99,235,0.35) !important;
}}
.stButton > button[kind="primary"]:hover {{
    box-shadow: 0 8px 24px rgba(37,99,235,0.5) !important;
    transform: translateY(-2px);
}}
.stDownloadButton > button {{
    background: #f9fafb !important;
    border: 1px solid #e5e7eb !important;
    color: #374151 !important;
    border-radius: 10px !important;
    font-size: 0.78rem !important;
    font-family: 'DM Sans', sans-serif !important;
}}
.stDownloadButton > button:hover {{
    background: linear-gradient(135deg, #0ea5e9, #2563eb) !important;
    color: #fff !important;
    border-color: transparent !important;
}}

/* ── Tabs ── */
.stTabs [data-baseweb="tab-list"] {{
    background: transparent;
    border-bottom: 1px solid #e5e7eb;
    gap: 2px;
}}
.stTabs [data-baseweb="tab"] {{
    background: #f9fafb !important;
    color: #6b7280 !important;
    border: 1px solid #e5e7eb !important;
    border-radius: 8px 8px 0 0 !important;
    padding: 0.5rem 1.1rem !important;
    font-weight: 500 !important;
    font-size: 0.82rem !important;
    font-family: 'DM Sans', sans-serif !important;
}}
.stTabs [aria-selected="true"] {{
    background: #ffffff !important;
    color: #0ea5e9 !important;
    border-bottom: 3px solid #0ea5e9 !important;
    font-weight: 600 !important;
}}

/* ── Metrics ── */
[data-testid="stMetric"] {{
    background: #ffffff;
    border: 1px solid #e5e7eb;
    border-radius: 16px;
    padding: 1rem 1.2rem;
    box-shadow: 0 4px 16px rgba(15,23,42,0.06);
}}
[data-testid="stMetricLabel"] {{ color: #6b7280 !important; font-size: 0.72rem !important; text-transform: uppercase; letter-spacing: 0.05em; }}
[data-testid="stMetricValue"] {{ font-size: 1.4rem !important; font-weight: 700 !important; color: #111827 !important; }}

/* ── Selectbox ── */
.stSelectbox [data-baseweb="select"] > div,
.stMultiSelect [data-baseweb="select"] > div {{
    background: #f9fafb !important;
    border-color: #e5e7eb !important;
    color: #111827 !important;
    border-radius: 10px !important;
}}

/* ── Expander ── */
.streamlit-expanderHeader {{
    background: #f9fafb !important;
    border-radius: 10px !important;
    color: #374151 !important;
    border: 1px solid #e5e7eb !important;
}}

/* ── Slider ── */
.stSlider [data-testid="stThumb"] {{ background: #0ea5e9 !important; }}
.stSlider [data-testid="stTrack"] > div:first-child {{ background: #0ea5e9 !important; }}

/* ── File uploader ── */
[data-testid="stFileUploadDropzone"] {{
    background: #f9fafb !important;
    border: 2px dashed #d1d5db !important;
    border-radius: 12px !important;
    color: #6b7280 !important;
}}
[data-testid="stFileUploadDropzone"]:hover {{ border-color: #0ea5e9 !important; }}

/* ── Checkbox ── */
.stCheckbox [data-testid="stWidgetLabel"] {{ color: #374151 !important; }}

/* ── Scrollbar ── */
::-webkit-scrollbar {{ width: 6px; height: 6px; }}
::-webkit-scrollbar-track {{ background: #f9fafb; }}
::-webkit-scrollbar-thumb {{ background: #d1d5db; border-radius: 3px; }}
::-webkit-scrollbar-thumb:hover {{ background: #0ea5e9; }}

/* ─────────────────────────────────────────────
   KPI ROW  — matches HTML .kpi-card
   ───────────────────────────────────────────── */
.kpi-row {{
    display: grid;
    gap: 0.75rem;
    margin: 0.8rem 0;
}}
.kpi-card {{
    background: #ffffff;
    border: 1px solid #e5e7eb;
    border-radius: 16px;
    padding: 1rem 1.1rem;
    box-shadow: 0 4px 16px rgba(15,23,42,0.06);
    transition: all 0.18s ease;
}}
.kpi-card:hover {{
    transform: translateY(-2px);
    box-shadow: 0 10px 30px rgba(15,23,42,0.10);
}}
.kpi-icon {{ font-size: 1.4rem; margin-bottom: 0.3rem; }}
.kpi-value {{ font-size: 1.3rem; font-weight: 700; line-height: 1.2; }}
.kpi-label {{ font-size: 0.68rem; text-transform: uppercase; letter-spacing: 0.12em; color: #9ca3af; margin-top: 0.2rem; }}
.kpi-meta {{ font-size: 0.72rem; color: #6b7280; margin-top: 0.1rem; }}

/* ─────────────────────────────────────────────
   HERO BANNER  — matches HTML .hero
   ───────────────────────────────────────────── */
.hero-banner {{
    background: #e8f3ff;
    border: 1px solid #bfdbfe;
    border-radius: 18px;
    padding: 1.3rem 1.6rem;
    margin-bottom: 1rem;
    box-shadow: 0 8px 30px rgba(15,23,42,0.08);
    position: relative;
    overflow: hidden;
    animation: fadeInUp 0.4s ease;
}}
.hero-banner::before {{
    content: '';
    position: absolute;
    right: -40px; top: -40px;
    width: 200px; height: 200px;
    border-radius: 999px;
    background: radial-gradient(circle, rgba(14,165,233,0.12) 0%, transparent 70%);
    pointer-events: none;
}}
.hero-heading {{
    font-size: 1.15rem;
    font-weight: 700;
    color: #0f172a;
    margin-bottom: 0.25rem;
}}
.hero-sub {{
    font-size: 0.82rem;
    color: #6b7280;
    margin-bottom: 0.7rem;
}}
.hero-tag-row {{ display: flex; flex-wrap: wrap; gap: 0.35rem; }}
.hero-tag {{
    font-size: 0.68rem; padding: 0.22rem 0.6rem; border-radius: 999px;
    border: 1px solid #bfdbfe; background: #eff6ff; color: #1d4ed8;
    font-weight: 500;
}}
.hero-status-badge {{
    display: inline-flex; align-items: center; gap: 0.35rem;
    padding: 0.3rem 0.7rem; border-radius: 999px;
    background: #ecfdf3; border: 1px solid #bbf7d0;
    font-size: 0.72rem; color: #166534; font-weight: 600;
}}
.hero-dot {{ display: inline-block; width: 8px; height: 8px; border-radius: 999px; background: #22c55e; }}
.hero-metric-val {{ font-size: 1.1rem; font-weight: 700; color: #0f172a; }}
.hero-metric-label {{ font-size: 0.68rem; color: #9ca3af; text-transform: uppercase; letter-spacing: 0.08em; }}

/* ─────────────────────────────────────────────
   SECTION HEADER
   ───────────────────────────────────────────── */
.section-header {{
    display: flex; align-items: center; gap: 0.6rem;
    margin: 1.2rem 0 0.6rem;
}}
.section-icon {{ font-size: 1.1rem; }}
.section-title {{ font-size: 0.92rem; font-weight: 600; color: #111827; }}
.section-sub {{ font-size: 0.74rem; color: #6b7280; margin-top: 2px; }}

/* ─────────────────────────────────────────────
   PANEL / CHART CARD
   ───────────────────────────────────────────── */
.panel {{
    background: #ffffff;
    border: 1px solid #e5e7eb;
    border-radius: 16px;
    padding: 1rem 1.1rem;
    box-shadow: 0 4px 16px rgba(15,23,42,0.06);
    margin: 0.5rem 0;
}}
.panel-title {{ font-size: 0.88rem; font-weight: 600; color: #111827; }}
.panel-sub {{ font-size: 0.74rem; color: #6b7280; margin-top: 2px; }}

/* ─────────────────────────────────────────────
   DATA TABLE  — matches HTML table styles
   ───────────────────────────────────────────── */
.table-wrap {{
    overflow-x: auto;
    border-radius: 12px;
    border: 1px solid #e5e7eb;
    margin: 0.5rem 0;
    box-shadow: 0 2px 8px rgba(15,23,42,0.04);
}}
table {{
    width: 100%;
    border-collapse: collapse;
    font-size: 0.78rem;
    font-family: 'DM Sans', sans-serif;
    background: #ffffff;
}}
th {{
    background: #f9fafb;
    text-align: left;
    font-size: 0.7rem;
    letter-spacing: 0.08em;
    text-transform: uppercase;
    color: #6b7280;
    padding: 0.55rem 0.7rem;
    border-bottom: 1px solid #e5e7eb;
    font-weight: 600;
    position: sticky; top: 0; z-index: 1;
}}
td {{
    padding: 0.48rem 0.7rem;
    border-bottom: 1px solid #f3f4f6;
    color: #374151;
    font-size: 0.78rem;
}}
tr.row-even td {{ background: #f9fafb; }}
tr.row-odd td {{ background: #ffffff; }}
tr.row-zero td {{ background: #fee2e2 !important; color: #7f1d1d; }}
tr:hover td {{ background: #eff6ff !important; }}

/* ─────────────────────────────────────────────
   PILLS  — exact match to HTML pill classes
   ───────────────────────────────────────────── */
.pill-danger {{
    font-size: 0.7rem; padding: 0.18rem 0.55rem; border-radius: 999px;
    border: 1px solid rgba(220,38,38,0.7); background: #fee2e2; color: #b91c1c;
    font-weight: 500;
}}
.pill-warning {{
    font-size: 0.7rem; padding: 0.18rem 0.55rem; border-radius: 999px;
    border: 1px solid rgba(245,158,11,0.7); background: #fef3c7; color: #92400e;
    font-weight: 500;
}}
.pill-ok {{
    font-size: 0.7rem; padding: 0.18rem 0.55rem; border-radius: 999px;
    border: 1px solid rgba(34,197,94,0.7); background: #dcfce7; color: #166534;
    font-weight: 500;
}}
.pill-soft {{
    font-size: 0.7rem; padding: 0.18rem 0.55rem; border-radius: 999px;
    border: 1px solid #e5e7eb; background: #f9fafb; color: #6b7280;
}}
.pill-currency {{
    font-size: 0.7rem; padding: 0.18rem 0.55rem; border-radius: 999px;
    border: 1px solid #bae6fd; background: #e0f2fe; color: #0369a1; font-weight: 600;
}}

/* ─────────────────────────────────────────────
   MINI CARDS  — matches HTML .mini-card
   ───────────────────────────────────────────── */
.mini-grid {{
    display: grid;
    grid-template-columns: repeat(2, minmax(0,1fr));
    gap: 0.6rem;
    margin-top: 0.5rem;
}}
.mini-card {{
    border-radius: 12px;
    background: #f9fafb;
    border: 1px solid #e5e7eb;
    padding: 0.65rem 0.75rem;
}}
.mini-label {{ font-size: 0.7rem; color: #9ca3af; text-transform: uppercase; letter-spacing: 0.08em; }}
.mini-value {{ font-size: 0.95rem; font-weight: 600; color: #111827; margin-top: 0.2rem; }}
.mini-meta {{ font-size: 0.68rem; color: #6b7280; margin-top: 0.15rem; }}

/* ─────────────────────────────────────────────
   EMPTY STATE
   ───────────────────────────────────────────── */
.empty-state {{
    text-align: center; padding: 2rem 1rem;
    color: #9ca3af; font-size: 0.85rem;
    background: #f9fafb; border-radius: 12px;
    border: 1px dashed #e5e7eb;
}}

/* ─────────────────────────────────────────────
   SYSTEM STATUS BADGES  — sidebar
   ───────────────────────────────────────────── */
.sys-badge {{
    display: inline-flex; align-items: center; gap: 5px;
    padding: 3px 10px; border-radius: 20px;
    font-size: 0.7rem; font-weight: 600;
    border: 1px solid transparent;
    margin: 2px;
}}
.sys-badge-ok {{ background: #ecfdf3; border-color: #bbf7d0; color: #166534; }}
.sys-badge-err {{ background: #fee2e2; border-color: #fecaca; color: #b91c1c; }}
.sys-badge-off {{ background: #f9fafb; border-color: #e5e7eb; color: #9ca3af; }}

/* ─────────────────────────────────────────────
   DIVIDER
   ───────────────────────────────────────────── */
.swag-divider {{
    border: none;
    border-top: 1px solid #e5e7eb;
    margin: 1.2rem 0;
}}

/* ─────────────────────────────────────────────
   LOGIN PAGE
   ───────────────────────────────────────────── */
.login-hero {{
    background: linear-gradient(135deg, #eff6ff 0%, #ecfdf5 100%);
    border: 1px solid #bfdbfe;
    border-radius: 20px;
    padding: 2.5rem;
    min-height: 480px;
    display: flex; flex-direction: column; justify-content: center;
    box-shadow: 0 8px 30px rgba(15,23,42,0.08);
}}
.login-form-card {{
    background: #ffffff;
    border: 1px solid #e5e7eb;
    border-radius: 20px;
    padding: 2.5rem;
    box-shadow: 0 8px 30px rgba(15,23,42,0.08);
}}

/* ─────────────────────────────────────────────
   ANIMATIONS
   ───────────────────────────────────────────── */
@keyframes fadeInUp {{
    from {{ opacity: 0; transform: translateY(16px); }}
    to   {{ opacity: 1; transform: translateY(0); }}
}}
@keyframes fadeInLeft {{
    from {{ opacity: 0; transform: translateX(-16px); }}
    to   {{ opacity: 1; transform: translateX(0); }}
}}
@keyframes pulseTitle {{
    0%,100% {{ opacity: 0.6; transform: scale(1); }}
    50%      {{ opacity: 0.9; transform: scale(1.04); }}
}}
@keyframes float {{
    0%,100% {{ transform: translateY(0); }}
    50%      {{ transform: translateY(-5px); }}
}}

/* ── Hide streamlit chrome ── */
footer {{ visibility: hidden; }}
#MainMenu {{ visibility: hidden; }}
.stDeployButton {{ display: none; }}

/* ── RTL table ── */
{"[dir='rtl'] th, [dir='rtl'] td { text-align: right; }" if is_rtl else ""}
</style>
"""

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 17: LOGIN PAGE
# ─────────────────────────────────────────────────────────────────────────────
def show_login():
    st.markdown(_build_css(), unsafe_allow_html=True)
    left, right = st.columns([1, 1], gap="large")

    with left:
        systems_html = "".join(
            f"<span class='sys-badge sys-badge-{'ok' if st.secrets.get(k,{}).get('url') else 'off'}'>"
            f"{'●' if st.secrets.get(k,{}).get('url') else '○'} {st.secrets.get(k,{}).get('name',k)}</span>"
            for k in SYSTEM_KEYS
        )
        chips = [
            t("Multi-System","متعدد الأنظمة"), t("Real-time","فوري"),
            t("Bilingual","ثنائي اللغة"), t("PDF Import","استيراد PDF"),
            t("Purchase Analytics","تحليل المشتريات"), t("Sales Analytics","تحليل المبيعات")
        ]
        chips_html = "".join(f"<span class='hero-tag'>{c}</span>" for c in chips)
        st.markdown(f"""
        <div class="login-hero" style="animation:fadeInLeft 0.5s ease;">
            <div style="font-size:3rem;margin-bottom:1rem;animation:float 3s ease-in-out infinite;">🏷️</div>
            <div style="font-size:2rem;font-weight:800;
                        background:linear-gradient(135deg,#0ea5e9,#22c55e);
                        -webkit-background-clip:text;-webkit-text-fill-color:transparent;
                        margin-bottom:0.4rem;">
                SWAG Control
            </div>
            <div style="color:#6b7280;font-size:0.9rem;margin-bottom:1.5rem;">
                {t('Product Comparison & Analytics v29','مقارنة المنتجات والتحليلات v29')}
            </div>
            <div class="hero-tag-row" style="margin-bottom:1.5rem;">{chips_html}</div>
            <div style="border-top:1px solid #e5e7eb;padding-top:1rem;margin-top:0.5rem;">
                <div style="font-size:0.7rem;color:#9ca3af;text-transform:uppercase;letter-spacing:0.1em;margin-bottom:0.6rem;">
                    {t('Connected Systems','الأنظمة المتصلة')}
                </div>
                <div style="display:flex;flex-wrap:wrap;gap:6px;">{systems_html}</div>
            </div>
        </div>
        """, unsafe_allow_html=True)

    with right:
        st.markdown(f"""
        <div class="login-form-card" style="animation:fadeInUp 0.5s ease;">
            <div style="font-size:1.5rem;font-weight:700;color:#0f172a;margin-bottom:0.3rem;">
                {t('Sign In','تسجيل الدخول')}
            </div>
            <div style="color:#6b7280;font-size:0.85rem;margin-bottom:1.5rem;">
                {t('Access your control center','الوصول إلى مركز التحكم')}
            </div>
        """, unsafe_allow_html=True)

        if st.session_state.get("_login_err"):
            st.markdown(f"""
            <div style="background:#fee2e2;border:1px solid #fecaca;border-radius:10px;
                        padding:10px 14px;color:#b91c1c;font-size:0.82rem;margin-bottom:1rem;">
                ❌ {st.session_state._login_err}
            </div>
            """, unsafe_allow_html=True)

        with st.form("login_form", clear_on_submit=False):
            email = st.text_input(t("Email", "البريد الإلكتروني"), placeholder="user@company.com")
            password = st.text_input(t("Password", "كلمة المرور"), type="password")
            submitted = st.form_submit_button(
                f"🚀 {t('Sign In','تسجيل الدخول')}", type="primary", use_container_width=True
            )
            if submitted:
                if not email.strip() or not password:
                    st.session_state._login_err = t("Please enter email and password.", "يرجى إدخال البريد الإلكتروني وكلمة المرور.")
                    st.rerun()
                else:
                    with st.spinner(t("Authenticating…", "جاري التحقق…")):
                        ok, err_msg = _attempt_login(email.strip(), password)
                    if ok:
                        st.session_state.authenticated = True
                        st.session_state.user_email = email.strip()
                        st.session_state._login_err = ""
                        try:
                            st.query_params.update({"u": email.strip(), "t": _make_token(email.strip())})
                        except Exception:
                            pass
                        st.balloons()
                        st.rerun()
                    else:
                        st.session_state._login_err = err_msg
                        st.rerun()

        st.markdown("</div>", unsafe_allow_html=True)

def _attempt_login(email, password):
    candidates = []
    if "LOGIN" in st.secrets:
        candidates.append(("LOGIN", st.secrets["LOGIN"]))
    for k in SYSTEM_KEYS:
        cfg = st.secrets.get(k, {})
        if cfg.get("url") and cfg.get("db"):
            candidates.append((k, cfg))
    if not candidates:
        return False, t("No Odoo connections configured.", "لا توجد اتصالات Odoo مُكوَّنة.")
    last_err = ""
    for src, cfg in candidates:
        url = cfg.get("url", "").rstrip("/")
        db = cfg.get("db", "")
        if not url or not db:
            continue
        try:
            proxy = xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/common", allow_none=True)
            uid = proxy.authenticate(db, email, password, {})
            if uid and isinstance(uid, int) and uid > 0:
                return True, ""
            last_err = t(f"Login failed on {db}.", f"فشل تسجيل الدخول على {db}.")
        except Exception as e:
            last_err = f"[{src}] {type(e).__name__}: {e}"
    return False, last_err

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 18: SIDEBAR  ── matches HTML sidebar design
# ─────────────────────────────────────────────────────────────────────────────
def _render_sidebar():
    with st.sidebar:
        # Brand
        st.markdown(f"""
        <div style="padding:0.8rem 0.2rem 1rem;border-bottom:1px solid #e5e7eb;margin-bottom:1rem;">
            <div style="display:flex;align-items:center;gap:0.6rem;">
                <div style="width:36px;height:36px;border-radius:12px;
                            background:linear-gradient(135deg,#0ea5e9,#22c55e);
                            display:flex;align-items:center;justify-content:center;
                            color:#fff;font-weight:700;font-size:0.9rem;
                            box-shadow:0 0 18px rgba(14,165,233,0.35);">S</div>
                <div>
                    <div style="font-size:0.92rem;font-weight:700;color:#111827;">SWAG Control</div>
                    <div style="font-size:0.68rem;color:#9ca3af;">Inventory Intelligence</div>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # User pill
        email = st.session_state.get("user_email", "")
        avatar = email[0].upper() if email else "U"
        st.markdown(f"""
        <div style="background:#f9fafb;border:1px solid #e5e7eb;border-radius:10px;
                    padding:0.55rem 0.75rem;margin-bottom:1rem;
                    display:flex;align-items:center;gap:0.6rem;">
            <div style="width:28px;height:28px;border-radius:999px;
                        background:linear-gradient(135deg,#0ea5e9,#22c55e);
                        display:flex;align-items:center;justify-content:center;
                        font-weight:700;font-size:0.8rem;color:#fff;">{avatar}</div>
            <div>
                <div style="font-size:0.75rem;color:#111827;font-weight:500;
                            white-space:nowrap;overflow:hidden;text-overflow:ellipsis;max-width:140px;">
                    {email}
                </div>
                <div style="font-size:0.63rem;color:#22c55e;font-weight:600;">● {t('Online','متصل')}</div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # Navigation
        st.markdown(f"<div style='font-size:0.68rem;color:#9ca3af;text-transform:uppercase;letter-spacing:0.1em;margin-bottom:6px;'>{t('Workspaces','الفضاءات')}</div>", unsafe_allow_html=True)
        nav_items = {
            "stock":    (f"📊 {t('Stock Overview','نظرة المخزون')}",    "SWAG"),
            "purchase": (f"🛒 {t('Purchase Analytics','تحليل المشتريات')}", "PO"),
            "sales":    (f"📈 {t('Sales Analytics','تحليل المبيعات')}",    "SO"),
        }
        current_view = st.session_state.get("analytics_view", "stock")
        for key, (label, meta) in nav_items.items():
            active = current_view == key
            if st.button(label, key=f"nav_{key}", use_container_width=True,
                         type="primary" if active else "secondary"):
                st.session_state.analytics_view = key
                st.rerun()

        st.markdown("<hr style='border-color:#e5e7eb;margin:0.8rem 0;'>", unsafe_allow_html=True)

        # Snapshots
        st.markdown(f"<div style='font-size:0.68rem;color:#9ca3af;text-transform:uppercase;letter-spacing:0.1em;margin-bottom:6px;'>{t('Snapshots','اللقطات')}</div>", unsafe_allow_html=True)

        total_df = st.session_state.get("total_df")
        reorder_df = st.session_state.get("reorder_df")
        total_qty = int(_to_num(total_df[C_ON_HAND]).sum()) if total_df is not None and not total_df.empty and C_ON_HAND in total_df.columns else 0
        reorder_cnt = len(reorder_df) if reorder_df is not None and not reorder_df.empty else 0
        crit_cnt = len(reorder_df[reorder_df[C_PRIORITY].str.contains("Critical", na=False)]) if reorder_df is not None and not reorder_df.empty and C_PRIORITY in reorder_df.columns else 0

        st.markdown(f"""
        <div style="background:#f9fafb;border:1px solid #e5e7eb;border-radius:12px;
                    padding:0.65rem 0.75rem;margin-bottom:0.5rem;font-size:0.72rem;">
            <div style="display:flex;justify-content:space-between;align-items:center;color:#111827;font-weight:600;">
                <span>Stock trend</span>
                <span style="color:#16a34a;">+12%</span>
            </div>
            <div style="font-size:0.68rem;color:#9ca3af;">Last 30 days</div>
            <div style="display:flex;gap:3px;margin-top:0.4rem;align-items:flex-end;">
                {"".join(f'<div style="width:6px;height:{h}px;border-radius:3px 3px 0 0;background:{c};"></div>'
                          for h,c in [(18,'#bfdbfe'),(26,'#60a5fa'),(14,'#bfdbfe'),(28,'#0ea5e9'),(20,'#bfdbfe'),(24,'#22c55e')])}
            </div>
        </div>

        <div style="background:#f9fafb;border:1px solid #e5e7eb;border-radius:12px;
                    padding:0.65rem 0.75rem;margin-bottom:0.5rem;font-size:0.72rem;">
            <div style="display:flex;justify-content:space-between;align-items:center;color:#111827;font-weight:600;">
                <span>Reorder risk</span>
                <span style="color:#dc2626;">{reorder_cnt} items</span>
            </div>
            <div style="font-size:0.68rem;color:#9ca3af;">Below threshold</div>
            <div style="display:flex;gap:4px;margin-top:0.45rem;flex-wrap:wrap;">
                <span class="pill-danger">Critical {crit_cnt}</span>
                <span class="pill-warning">Low</span>
                <span class="pill-ok">OK</span>
            </div>
        </div>

        <div style="background:#f9fafb;border:1px solid #e5e7eb;border-radius:12px;
                    padding:0.65rem 0.75rem;margin-bottom:0.5rem;font-size:0.72rem;">
            <div style="display:flex;justify-content:space-between;align-items:center;color:#111827;font-weight:600;">
                <span>Sales vs stock</span>
                <span style="color:#16a34a;">72%</span>
            </div>
            <div style="font-size:0.68rem;color:#9ca3af;">Sell-through</div>
            <div style="width:100%;height:8px;border-radius:999px;background:#e5e7eb;margin-top:0.5rem;overflow:hidden;">
                <div style="width:72%;height:100%;background:linear-gradient(90deg,#22c55e,#0ea5e9);"></div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("<hr style='border-color:#e5e7eb;margin:0.8rem 0;'>", unsafe_allow_html=True)

        # Search settings
        st.markdown(f"<div style='font-size:0.68rem;color:#9ca3af;text-transform:uppercase;letter-spacing:0.1em;margin-bottom:6px;'>{t('Search Settings','إعدادات البحث')}</div>", unsafe_allow_html=True)
        st.session_state.low_stock_thresh = st.slider(
            t("Low Stock Threshold", "حد المخزون المنخفض"), 0, 50, st.session_state.low_stock_thresh)
        st.session_state.search_exact = st.checkbox(
            t("Exact Match", "تطابق تام"), value=st.session_state.search_exact)
        st.session_state.show_transfers = st.checkbox(
            t("Show Transfers", "عرض التحويلات"), value=st.session_state.show_transfers)
        st.session_state.show_reorder = st.checkbox(
            t("Show Reorder", "عرض قائمة الطلب"), value=st.session_state.show_reorder)

        with st.expander(f"⚙️ {t('Reorder Settings','إعدادات إعادة الطلب')}"):
            st.session_state.reorder_mode = st.selectbox(
                t("Mode", "النمط"), ["days_cover", "reorder_point"],
                index=0 if st.session_state.reorder_mode == "days_cover" else 1)
            st.session_state.reorder_target_days = st.number_input(
                t("Target Days", "أيام التغطية"), min_value=1, max_value=365,
                value=st.session_state.reorder_target_days)
            st.session_state.reorder_max_level = st.number_input(
                t("Max Level", "الحد الأقصى"), min_value=0, value=st.session_state.reorder_max_level)
            st.session_state.reorder_point = st.number_input(
                t("Reorder Point", "نقطة الطلب"), min_value=0, value=st.session_state.reorder_point)

        st.markdown("<hr style='border-color:#e5e7eb;margin:0.8rem 0;'>", unsafe_allow_html=True)

        # System status
        st.markdown(f"<div style='font-size:0.68rem;color:#9ca3af;text-transform:uppercase;letter-spacing:0.1em;margin-bottom:6px;'>{t('System Status','حالة الأنظمة')}</div>", unsafe_allow_html=True)
        sys_stats = st.session_state.get("sys_stats", {})
        for k in SYSTEM_KEYS:
            cfg = st.secrets.get(k, {})
            name = cfg.get("name_ar" if get_lang() == "AR" else "name", k)
            stat = sys_stats.get(k, {})
            level = stat.get("level", "off") if stat else ("ok" if cfg.get("url") else "off")
            icon = "✅" if level == "ok" else ("❌" if level == "error" else "⚫")
            cls = "ok" if level == "ok" else ("err" if level == "error" else "off")
            st.markdown(f"<span class='sys-badge sys-badge-{cls}'>{icon} {name}</span><br>", unsafe_allow_html=True)

        st.markdown("<hr style='border-color:#e5e7eb;margin:0.8rem 0;'>", unsafe_allow_html=True)

        # Language
        lang_options = ["EN", "AR"]
        new_lang = st.radio(f"🌐 {t('Language','اللغة')}", lang_options,
                            index=lang_options.index(get_lang()), horizontal=True)
        if new_lang != get_lang():
            st.session_state.lang = new_lang
            st.rerun()

        if st.button(f"🚪 {t('Logout','تسجيل الخروج')}", use_container_width=True):
            do_logout()

        # Footer card
        st.markdown(f"""
        <div style="margin-top:1rem;padding:0.75rem;border-radius:12px;
                    background:#eff6ff;border:1px solid #dbeafe;font-size:0.74rem;color:#6b7280;">
            <div style="font-size:0.78rem;font-weight:600;color:#0f172a;margin-bottom:0.2rem;">
                SWAG connected
            </div>
            {t('Dashboard wired to SWAG Odoo backend.','متصل بخلفية Odoo SWAG.')}
        </div>
        """, unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 19: STOCK COMPARISON VIEW
# ─────────────────────────────────────────────────────────────────────────────
def show_stock_comparison():
    total_df = st.session_state.get("total_df")
    sys_stats = st.session_state.get("sys_stats", {})
    systems_online = sum(1 for v in sys_stats.values() if v.get("level") == "ok")
    total_units = int(_to_num(total_df[C_ON_HAND]).sum()) if total_df is not None and not total_df.empty and C_ON_HAND in total_df.columns else 0
    models_found = total_df[C_MODEL].nunique() if total_df is not None and not total_df.empty else 0

    # ── Topbar row ─────────────────────────────────────────────────────────────
    today_str = datetime.now().strftime("%d %b %Y  %H:%M")
    col_top1, col_top2 = st.columns([2, 1])
    with col_top1:
        st.markdown(f"<div style='font-size:0.72rem;letter-spacing:0.14em;text-transform:uppercase;color:#9ca3af;margin-bottom:0.5rem;'>Executive overview</div>", unsafe_allow_html=True)
    with col_top2:
        st.markdown(f"<div style='text-align:right;font-size:0.72rem;color:#9ca3af;margin-bottom:0.5rem;'>🕐 {today_str}</div>", unsafe_allow_html=True)

    # ── Hero banner ─────────────────────────────────────────────────────────────
    sys_badges = "".join(
        f"<span class='sys-badge sys-badge-{'ok' if v.get('level')=='ok' else 'err'}'>"
        f"{'✅' if v.get('level')=='ok' else '❌'} {v.get('system', k)}</span> "
        for k, v in sys_stats.items()
    ) if sys_stats else ""

    col_hero_l, col_hero_r = st.columns([3, 1])
    with col_hero_l:
        st.markdown(f"""
        <div class="hero-banner">
            <div class="hero-heading">📊 {t('SWAG Stock Overview','نظرة المخزون SWAG')}</div>
            <div class="hero-sub">{t('Live picture of SWAG inventory health across all branches.','صورة حية لصحة مخزون SWAG عبر جميع الفروع.')}</div>
            <div class="hero-tag-row">
                <span class="hero-tag">SWAG ERP</span>
                <span class="hero-tag">{t('Inventory','المخزون')}</span>
                <span class="hero-tag">{t('Real-time','فوري')}</span>
                {sys_badges}
            </div>
        </div>
        """, unsafe_allow_html=True)
    with col_hero_r:
        st.markdown(f"""
        <div style="background:#ffffff;border:1px solid #e5e7eb;border-radius:16px;
                    padding:1rem;box-shadow:0 4px 16px rgba(15,23,42,0.06);height:100%;
                    display:flex;flex-direction:column;justify-content:center;align-items:center;gap:0.8rem;">
            <div class="hero-status-badge"><span class="hero-dot"></span> {t('SWAG connected','SWAG متصل')}</div>
            <div style="display:flex;gap:1.5rem;">
                <div style="text-align:center;">
                    <div class="hero-metric-val">{total_units:,}</div>
                    <div class="hero-metric-label">{t('Units','وحدات')}</div>
                </div>
                <div style="text-align:center;">
                    <div class="hero-metric-val">{models_found:,}</div>
                    <div class="hero-metric-label">{t('Models','موديلات')}</div>
                </div>
                <div style="text-align:center;">
                    <div class="hero-metric-val">{systems_online}/4</div>
                    <div class="hero-metric-label">{t('Online','متصل')}</div>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)

    # ── PDF Upload ──────────────────────────────────────────────────────────────
    _section_header(t("Import Product Codes", "استيراد رموز المنتجات"), "📄",
                    t("Upload a PDF invoice to extract model codes", "ارفع فاتورة PDF لاستخراج رموز الموديلات"))

    with st.expander(t("📄 PDF Invoice Upload", "📄 رفع فاتورة PDF"), expanded=st.session_state.get("pdf_mode", False)):
        uploaded = st.file_uploader(t("Upload PDF", "رفع PDF"), type=["pdf"], key="pdf_uploader")
        if uploaded:
            with st.spinner(t("Parsing invoice…", "جاري تحليل الفاتورة…")):
                raw = parse_invoice_pdf_cached(uploaded.read())
            base_models = get_unique_base_models(raw)
            if base_models:
                st.success(t(f"Found {len(base_models)} unique base models.", f"تم العثور على {len(base_models)} موديل."))
                st.session_state.pdf_codes = base_models
                st.session_state.pdf_mode = True
                with st.expander(t("Preview Codes", "معاينة الرموز")):
                    st.code(", ".join(base_models))
            else:
                st.warning(t("No model codes found in PDF.", "لم يتم العثور على رموز في PDF."))
        if st.session_state.get("pdf_codes"):
            st.markdown(f"<div class='pill-ok' style='display:inline-block;'>✅ {len(st.session_state.pdf_codes)} {t('codes loaded from PDF','رمز محمل من PDF')}</div>", unsafe_allow_html=True)
            if st.button(t("Clear PDF Codes", "مسح رموز PDF"), key="clear_pdf"):
                st.session_state.pdf_codes = []
                st.session_state.pdf_mode = False
                st.rerun()

    # ── Search ──────────────────────────────────────────────────────────────────
    _section_header(t("Product Search & Filters", "بحث المنتجات والفلاتر"), "🔍",
                    t("Search SWAG models and review stock / reorder suggestions","ابحث عن موديلات SWAG وراجع المخزون"))

    st.markdown("<div class='panel'>", unsafe_allow_html=True)
    pdf_pre = ", ".join(st.session_state.get("pdf_codes", []))
    code_input = st.text_area(
        t("Model codes (comma or newline separated)", "رموز الموديلات (مفصولة بفاصلة أو سطر جديد)"),
        value=pdf_pre, height=80,
        placeholder="e.g. ABC-001, DEF-002, GHI-003",
        key="code_input_area",
    )
    col_btn, col_opt = st.columns([2, 1])
    with col_btn:
        search_clicked = st.button(
            f"🔍 {t('Run Search','تشغيل البحث')}", type="primary",
            use_container_width=True, key="search_btn"
        )
    with col_opt:
        st.session_state.search_exact = st.checkbox(
            t("Exact Match", "تطابق تام"), value=st.session_state.search_exact, key="exact_cb"
        )
    st.markdown("</div>", unsafe_allow_html=True)

    if search_clicked:
        raw_codes = [c.strip().upper() for c in re.split(r"[,\n;]+", code_input) if c.strip()]
        codes = list(dict.fromkeys(raw_codes))
        with st.spinner(t("Fetching data from all systems…", "جاري جلب البيانات…")):
            total_df, branch_df, transfers_df, reorder_df, sys_stats = fetch_all_data(
                codes, st.session_state.search_exact, st.session_state.low_stock_thresh,
                st.session_state.show_transfers, st.session_state.show_reorder,
                st.session_state.reorder_mode, st.session_state.reorder_target_days,
                st.session_state.reorder_max_level, st.session_state.reorder_point,
            )
        st.session_state.total_df = total_df
        st.session_state.branch_df = branch_df
        st.session_state.transfers_df = transfers_df
        st.session_state.reorder_df = reorder_df
        st.session_state.sys_stats = sys_stats
        st.session_state.last_run = datetime.now()
        if total_df is not None and not total_df.empty:
            price_history = {}
            for _, row in total_df.iterrows():
                mc = row.get(C_MODEL, ""); sys = row.get(C_SYSTEM, ""); price = float(row.get(C_SALE_PRICE, 0) or 0)
                if mc and sys:
                    if mc not in price_history: price_history[mc] = {}
                    price_history[mc][sys] = price
            st.session_state.price_history = price_history
        for pk in ["page_total","page_branch","page_transfers","page_reorder"]:
            st.session_state[pk] = 0
        st.rerun()

    # ── Results ─────────────────────────────────────────────────────────────────
    total_df = st.session_state.get("total_df")
    if total_df is None or total_df.empty:
        if st.session_state.get("last_run") is not None:
            st.markdown(f"<div class='empty-state'>ℹ️ {t('No data returned.','لا توجد بيانات.')}</div>", unsafe_allow_html=True)
        return

    _divider()

    branch_df = st.session_state.get("branch_df")
    transfers_df = st.session_state.get("transfers_df")
    reorder_df = st.session_state.get("reorder_df")

    # ── KPI row ─────────────────────────────────────────────────────────────────
    total_qty = int(_to_num(total_df[C_ON_HAND]).sum())
    total_val = (_to_num(total_df[C_ON_HAND]) * _to_num(total_df[C_SALE_PRICE])).sum()
    zero_cnt = int((_to_num(total_df[C_ON_HAND]) == 0).sum())
    low_cnt  = int(((_to_num(total_df[C_ON_HAND]) > 0) & (_to_num(total_df[C_ON_HAND]) <= st.session_state.low_stock_thresh)).sum())

    _render_kpi_grid([
        _kpi_card("📦", f"{total_qty:,}", t("Total On Hand","إجمالي المتوفر"), "SWAG inventory", "#0ea5e9"),
        _kpi_card("💰", f"SAR {total_val:,.0f}", t("Stock Value","قيمة المخزون"), "Qty × price", "#22c55e"),
        _kpi_card("🔴", f"{zero_cnt:,}", t("Zero Stock","صفر مخزون"), "Need purchase/transfer", "#ef4444"),
        _kpi_card("⚠️", f"{low_cnt:,}", t("Low Stock","مخزون منخفض"), f"Below {st.session_state.low_stock_thresh}", "#f59e0b"),
    ])

    # ── Charts ──────────────────────────────────────────────────────────────────
    _section_header(t("Visualizations","المرئيات"), "📊")
    col_c1, col_c2 = st.columns(2)
    with col_c1:
        st.markdown("<div class='panel'>", unsafe_allow_html=True)
        st.markdown("<div class='panel-title'>Category wise stock</div>", unsafe_allow_html=True)
        st.markdown("<div class='panel-sub'>On-hand quantity by category.</div>", unsafe_allow_html=True)
        if C_CATEGORY in total_df.columns:
            cat_agg = total_df.groupby(C_CATEGORY)[C_ON_HAND].sum().reset_index().nlargest(8, C_ON_HAND)
            fig = px.bar(cat_agg, x=C_CATEGORY, y=C_ON_HAND,
                         color=C_ON_HAND, color_continuous_scale=["#bfdbfe","#3b82f6"],
                         template="plotly_white")
            fig.update_traces(marker_cornerradius=5, marker_line_width=0)
            st.plotly_chart(_apply_light_theme(fig), use_container_width=True)
        st.markdown("</div>", unsafe_allow_html=True)

    with col_c2:
        st.markdown("<div class='panel'>", unsafe_allow_html=True)
        st.markdown("<div class='panel-title'>Brand share</div>", unsafe_allow_html=True)
        st.markdown("<div class='panel-sub'>Stock distribution by system.</div>", unsafe_allow_html=True)
        if C_SYSTEM in total_df.columns:
            sys_agg = total_df.groupby(C_SYSTEM)[C_ON_HAND].sum().reset_index()
            fig = _donut_chart(sys_agg, C_SYSTEM, C_ON_HAND, "")
            if fig:
                st.plotly_chart(fig, use_container_width=True)
        st.markdown("</div>", unsafe_allow_html=True)

    col_c3, col_c4 = st.columns(2)
    with col_c3:
        st.markdown("<div class='panel'>", unsafe_allow_html=True)
        st.markdown("<div class='panel-title'>Top 10 models by qty</div>", unsafe_allow_html=True)
        st.markdown("<div class='panel-sub'>Highest on-hand stock models.</div>", unsafe_allow_html=True)
        fig = _top10_bar(total_df, C_MODEL, C_ON_HAND, "", "#0ea5e9")
        if fig:
            st.plotly_chart(fig, use_container_width=True)
        st.markdown("</div>", unsafe_allow_html=True)

    with col_c4:
        st.markdown("<div class='panel'>", unsafe_allow_html=True)
        st.markdown("<div class='panel-title'>Price comparison by system</div>", unsafe_allow_html=True)
        st.markdown("<div class='panel-sub'>Sale price comparison across systems.</div>", unsafe_allow_html=True)
        price_history = st.session_state.get("price_history", {})
        if price_history:
            price_rows = [{"Model": mc, "System": sys, "Price": price}
                          for mc, sys_prices in price_history.items()
                          for sys, price in sys_prices.items()]
            if price_rows:
                price_df = pd.DataFrame(price_rows)
                fig_p = px.line(price_df, x="Model", y="Price", color="System",
                                markers=True, color_discrete_sequence=_COLORS, template="plotly_white")
                st.plotly_chart(_apply_light_theme(fig_p), use_container_width=True)
        else:
            st.markdown("<div class='empty-state'>Run a search to see price comparison.</div>", unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)

    # ── Data table tabs ──────────────────────────────────────────────────────────
    _section_header(t("Product Search & Detail","تفاصيل المنتجات"), "📋",
                    t("Search SWAG products, low stock, and reorder suggestions","ابحث عن منتجات SWAG والمخزون المنخفض"))

    tab_labels = [f"📊 {t('Total Stock','إجمالي المخزون')}",
                  f"🏪 {t('By Branch','حسب الفرع')}"]
    if st.session_state.show_transfers:
        tab_labels.append(f"🔄 {t('Transfers','التحويلات')}")
    if st.session_state.show_reorder:
        tab_labels.append(f"🔔 {t('Reorder Risk','مخاطر الطلب')}")

    tabs = st.tabs(tab_labels)
    tab_idx = 0

    # Total stock
    with tabs[tab_idx]:
        tab_idx += 1
        col_left, col_right = st.columns([3, 1])
        with col_left:
            st.markdown(f"<div class='panel-title'>📊 {t('Total Stock','إجمالي المخزون')}</div>", unsafe_allow_html=True)
            st.markdown(f"<div class='panel-sub'>{t('Loaded','محمل')} {len(total_df)} {t('SWAG product rows.','سطر منتج SWAG.')}</div>", unsafe_allow_html=True)
        with col_right:
            _render_mini_grid([
                _mini_card(t("Zero-stock models","موديلات بدون مخزون"), str(zero_cnt), t("Need action","تحتاج إجراء"), "#ef4444"),
                _mini_card(t("Tracked models","الموديلات المتتبعة"), str(total_df[C_MODEL].nunique()), t("In current view","في العرض الحالي")),
            ])

        display_df(total_df, st.session_state.low_stock_thresh, "total_main")
        ec1, ec2 = st.columns(2)
        with ec1:
            st.download_button(f"⬇️ {t('Export CSV','CSV')}", to_csv(df_for_display(total_df)), dl_name("total_stock","csv"), "text/csv", use_container_width=True)
        with ec2:
            st.download_button(f"⬇️ {t('Export Excel','Excel')}", to_excel(df_for_display(total_df)), dl_name("total_stock","xlsx"), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

    # Branch stock
    with tabs[tab_idx]:
        tab_idx += 1
        if branch_df is None or branch_df.empty:
            st.markdown(f"<div class='empty-state'>ℹ️ {t('No branch data.','لا توجد بيانات فروع.')}</div>", unsafe_allow_html=True)
        else:
            br_agg = branch_df.groupby(C_BRANCH)[C_ON_HAND].sum().reset_index().nlargest(15, C_ON_HAND)
            fig_br = px.bar(br_agg, x=C_BRANCH, y=C_ON_HAND,
                            color=C_ON_HAND, color_continuous_scale=["#bfdbfe","#0ea5e9"],
                            template="plotly_white")
            fig_br.update_traces(marker_cornerradius=5, marker_line_width=0)
            st.plotly_chart(_apply_light_theme(fig_br), use_container_width=True)

            perBranch = branch_df.groupby(C_BRANCH)[C_ON_HAND].sum().reset_index().sort_values(C_ON_HAND, ascending=False)
            top_branch = perBranch.iloc[0][C_BRANCH] if len(perBranch) > 0 else "—"
            _render_mini_grid([
                _mini_card(t("Top branch","أفضل فرع"), str(top_branch), f"Qty {int(perBranch.iloc[0][C_ON_HAND]) if len(perBranch)>0 else 0}"),
                _mini_card(t("Branches","الفروع"), str(len(perBranch)), t("Active in result","نشطة في النتائج")),
            ])

            display_df(branch_df, st.session_state.low_stock_thresh, "branch_main")
            bc1, bc2, bc3 = st.columns(3)
            with bc1:
                st.download_button(f"⬇️ CSV", to_csv(df_for_display(branch_df)), dl_name("branch","csv"), "text/csv", use_container_width=True)
            with bc2:
                st.download_button(f"⬇️ Excel", to_excel(df_for_display(branch_df)), dl_name("branch","xlsx"), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
            with bc3:
                try:
                    pivot = branch_df.pivot_table(index=C_MODEL, columns=C_BRANCH, values=C_ON_HAND, aggfunc="sum", fill_value=0)
                    out = io.BytesIO()
                    with pd.ExcelWriter(out, engine="openpyxl") as w:
                        pivot.to_excel(w, sheet_name="Branch_Matrix")
                    out.seek(0)
                    st.download_button(f"📊 {t('Branch Matrix','مصفوفة')}", out, dl_name("matrix","xlsx"), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
                except Exception:
                    pass

    # Transfers
    if st.session_state.show_transfers and tab_idx < len(tabs):
        with tabs[tab_idx]:
            tab_idx += 1
            if transfers_df is None or transfers_df.empty:
                st.markdown(f"<div class='empty-state'>ℹ️ {t('No pending transfers.','لا توجد تحويلات معلقة.')}</div>", unsafe_allow_html=True)
            else:
                tf_total = int(_to_num(transfers_df[C_QTY]).sum()) if C_QTY in transfers_df.columns else 0
                _render_mini_grid([
                    _mini_card(t("Transfer Lines","سطور التحويل"), f"{len(transfers_df):,}", t("Pending","معلق")),
                    _mini_card(t("Total Qty Moving","الكمية المتحركة"), f"{tf_total:,}", t("In transit","في الطريق")),
                ])
                display_df(transfers_df, 0, "transfers_main")
                st.download_button(f"⬇️ CSV", to_csv(df_for_display(transfers_df)), dl_name("transfers","csv"), "text/csv")

    # Reorder
    if st.session_state.show_reorder and tab_idx < len(tabs):
        with tabs[tab_idx]:
            if reorder_df is None or reorder_df.empty:
                st.markdown(f"<div style='text-align:center;padding:1.5rem;'><span class='pill-ok'>✅ {t('No reorder needed.','لا حاجة لإعادة الطلب.')}</span></div>", unsafe_allow_html=True)
            else:
                crit_cnt = len(reorder_df[reorder_df[C_PRIORITY].str.contains("Critical", na=False)]) if C_PRIORITY in reorder_df.columns else 0
                _render_mini_grid([
                    _mini_card(t("Items to Reorder","عناصر للطلب"), f"{len(reorder_df):,}", t("Based on rule","بناءً على القاعدة")),
                    _mini_card(t("Critical Priority","أولوية حرجة"), str(crit_cnt), t("Immediate PO suggestion","اقتراح أمر شراء فوري"), "#ef4444"),
                ])
                if C_PRIORITY in reorder_df.columns:
                    crit_df = reorder_df[reorder_df[C_PRIORITY].str.contains("Critical", na=False)]
                    if not crit_df.empty:
                        st.markdown(f"<div class='pill-danger' style='display:inline-block;margin-bottom:8px;'>🔴 {t('Critical Items — Action Required','عناصر حرجة')}</div>", unsafe_allow_html=True)
                        display_df(crit_df, 0, "reorder_crit")
                display_df(reorder_df, 0, "reorder_main")
                rc1, rc2 = st.columns(2)
                with rc1:
                    st.download_button(f"⬇️ CSV", to_csv(df_for_display(reorder_df)), dl_name("reorder","csv"), "text/csv", use_container_width=True)
                with rc2:
                    st.download_button(f"⬇️ Excel", to_excel(df_for_display(reorder_df)), dl_name("reorder","xlsx"), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 20: PURCHASE ANALYTICS VIEW
# ─────────────────────────────────────────────────────────────────────────────
def show_purchase_analytics():
    today_str = datetime.now().strftime("%d %b %Y  %H:%M")
    st.markdown(f"<div style='font-size:0.72rem;letter-spacing:0.14em;text-transform:uppercase;color:#9ca3af;margin-bottom:0.5rem;'>Executive overview &nbsp;|&nbsp; {today_str}</div>", unsafe_allow_html=True)

    st.markdown(f"""
    <div class="hero-banner">
        <div class="hero-heading">🛒 {t('SWAG Purchase Analytics','تحليل مشتريات SWAG')}</div>
        <div class="hero-sub">{t('See how much you are buying, from whom, and days of coverage per PO.','شاهد ما تشتريه وممن ومدة التغطية لكل أمر شراء.')}</div>
        <div class="hero-tag-row">
            <span class="hero-tag">Purchase Orders</span>
            <span class="hero-tag">Vendor Analysis</span>
            <span class="hero-tag">Days Cover</span>
        </div>
    </div>
    """, unsafe_allow_html=True)

    _section_header(t("Filters","الفلاتر"), "⚙️")
    st.markdown("<div class='panel'>", unsafe_allow_html=True)
    col_d1, col_d2, col_mc = st.columns([1, 1, 2])
    with col_d1:
        date_from = st.date_input(t("From Date","من تاريخ"), value=datetime.now()-timedelta(days=30), key="po_from")
    with col_d2:
        date_to = st.date_input(t("To Date","إلى تاريخ"), value=datetime.now(), key="po_to")
    with col_mc:
        mc_input = st.text_input(t("Model Code Filter (optional)","فلتر رمز الموديل"), placeholder="e.g. ABC-001", key="po_mc_filter")
    if st.button(f"🔍 {t('Fetch Purchase Data','جلب بيانات المشتريات')}", type="primary", key="po_fetch", use_container_width=True):
        mc_list = [c.strip().upper() for c in re.split(r"[,\n;]+", mc_input) if c.strip()] if mc_input.strip() else []
        with st.spinner(t("Fetching purchase history…","جاري جلب سجل المشتريات…")):
            po_df = fetch_all_systems_purchase_history(mc_list, str(date_from), str(date_to))
        st.session_state.po_analytics_df = po_df
        st.session_state.page_po = 0
        st.rerun()
    st.markdown("</div>", unsafe_allow_html=True)

    po_df = st.session_state.get("po_analytics_df")
    if po_df is None or po_df.empty:
        if st.session_state.po_analytics_df is not None:
            st.markdown(f"<div class='empty-state'>ℹ️ {t('No purchase data found.','لا توجد بيانات مشتريات.')}</div>", unsafe_allow_html=True)
        return

    _divider()

    # KPIs
    total_spend = _to_num(po_df[C_SUBTOTAL]).sum() if C_SUBTOTAL in po_df.columns else 0
    total_qty_p = _to_num(po_df[C_QTY_PURCHASED]).sum() if C_QTY_PURCHASED in po_df.columns else 0
    unique_pos = po_df[C_PO].nunique() if C_PO in po_df.columns else 0
    top_vendor = po_df.groupby(C_VENDOR)[C_SUBTOTAL].sum().idxmax() if C_VENDOR in po_df.columns and not po_df.empty else "—"

    _render_kpi_grid([
        _kpi_card("💰", f"SAR {total_spend:,.0f}", t("Total Spend","إجمالي الإنفاق"), "All PO lines", "#0ea5e9"),
        _kpi_card("📦", f"{total_qty_p:,.0f}", t("Units Bought","الوحدات المشتراة"), "All systems", "#22c55e"),
        _kpi_card("📋", f"{unique_pos:,}", t("Purchase Orders","أوامر الشراء"), "Selected period", "#f59e0b"),
        _kpi_card("🏭", str(top_vendor)[:16], t("Top Vendor","أفضل مورد"), "By spend", "#a855f7"),
    ])

    # Charts
    col_c1, col_c2 = st.columns(2)
    with col_c1:
        st.markdown("<div class='panel'>", unsafe_allow_html=True)
        st.markdown("<div class='panel-title'>Top Vendors</div><div class='panel-sub'>Purchase subtotal by vendor.</div>", unsafe_allow_html=True)
        if C_VENDOR in po_df.columns:
            vendor_agg = po_df.groupby(C_VENDOR)[C_SUBTOTAL].sum().reset_index().nlargest(6, C_SUBTOTAL)
            fig = px.bar(vendor_agg, x=C_SUBTOTAL, y=C_VENDOR, orientation="h",
                         color=C_SUBTOTAL, color_continuous_scale=["#bfdbfe","#0ea5e9"],
                         template="plotly_white")
            fig.update_traces(marker_cornerradius=5, marker_line_width=0)
            st.plotly_chart(_apply_light_theme(fig), use_container_width=True)
        st.markdown("</div>", unsafe_allow_html=True)

    with col_c2:
        st.markdown("<div class='panel'>", unsafe_allow_html=True)
        st.markdown("<div class='panel-title'>Daily Purchase Spend</div><div class='panel-sub'>Date wise purchase amount.</div>", unsafe_allow_html=True)
        if C_DATE in po_df.columns:
            daily = po_df.copy()
            daily["_date"] = pd.to_datetime(daily[C_DATE], errors="coerce").dt.date
            daily_agg = daily.groupby("_date")[C_SUBTOTAL].sum().reset_index()
            daily_agg.columns = ["Date","Spend"]
            if not daily_agg.empty:
                fig2 = px.area(daily_agg, x="Date", y="Spend",
                               color_discrete_sequence=["#0ea5e9"], template="plotly_white")
                fig2.update_traces(fill="tozeroy", fillcolor="rgba(14,165,233,0.10)")
                st.plotly_chart(_apply_light_theme(fig2), use_container_width=True)
        st.markdown("</div>", unsafe_allow_html=True)

    col_c3, col_c4 = st.columns(2)
    with col_c3:
        st.markdown("<div class='panel'>", unsafe_allow_html=True)
        fig = _top10_bar(po_df, C_MODEL, C_SUBTOTAL, "Top Products by Spend", "#22c55e")
        if fig:
            st.plotly_chart(fig, use_container_width=True)
        st.markdown("</div>", unsafe_allow_html=True)
    with col_c4:
        st.markdown("<div class='panel'>", unsafe_allow_html=True)
        if C_CATEGORY in po_df.columns:
            cat_agg = po_df.groupby(C_CATEGORY)[C_SUBTOTAL].sum().reset_index().nlargest(8, C_SUBTOTAL)
            fig_cat = _donut_chart(cat_agg, C_CATEGORY, C_SUBTOTAL, "Spend by Category")
            if fig_cat:
                st.plotly_chart(fig_cat, use_container_width=True)
        st.markdown("</div>", unsafe_allow_html=True)

    # Table
    _section_header(t("Purchase Detail","تفاصيل المشتريات"), "📋",
                    t("PO data with stock estimates via SWAG backend.","بيانات أوامر الشراء."))
    display_df(po_df, 0, "po_table")
    pc1, pc2 = st.columns(2)
    with pc1:
        st.download_button(f"⬇️ {t('Export CSV','CSV')}", to_csv(df_for_display(po_df)), dl_name("purchase","csv"), "text/csv", use_container_width=True)
    with pc2:
        st.download_button(f"⬇️ {t('Export Excel','Excel')}", to_excel(df_for_display(po_df)), dl_name("purchase","xlsx"), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

    # Per-code
    _section_header(t("Per-Code Analysis","تحليل لكل رمز"), "🔎")
    codes_in_df = sorted(po_df[C_MODEL].dropna().unique().tolist()) if C_MODEL in po_df.columns else []
    if codes_in_df:
        selected_code = st.selectbox(t("Select Model Code","اختر رمز الموديل"), codes_in_df, key="po_code_select")
        if selected_code:
            code_po = po_df[po_df[C_MODEL] == selected_code]
            total_df_cur = st.session_state.get("total_df")
            code_stock = total_df_cur[total_df_cur[C_MODEL] == selected_code] if total_df_cur is not None else None
            dc1, dc2 = st.columns(2)
            with dc1:
                st.markdown("<div class='panel'>", unsafe_allow_html=True)
                st.markdown(f"<div class='panel-title'>📋 {t('Purchase History','سجل المشتريات')}</div>", unsafe_allow_html=True)
                render_premium_table(df_for_display(code_po))
                st.markdown("</div>", unsafe_allow_html=True)
            with dc2:
                st.markdown("<div class='panel'>", unsafe_allow_html=True)
                st.markdown(f"<div class='panel-title'>📦 {t('Current Stock','المخزون الحالي')}</div>", unsafe_allow_html=True)
                if code_stock is not None and not code_stock.empty:
                    render_premium_table(df_for_display(code_stock[[C_SYSTEM,C_PRODUCT,C_ON_HAND,C_SALE_PRICE]]))
                else:
                    st.markdown(f"<div class='empty-state'>{t('No stock data. Run a stock search first.','لا توجد بيانات مخزون.')}</div>", unsafe_allow_html=True)
                st.markdown("</div>", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 21: SALES ANALYTICS VIEW
# ─────────────────────────────────────────────────────────────────────────────
def show_sales_analytics():
    today_str = datetime.now().strftime("%d %b %Y  %H:%M")
    st.markdown(f"<div style='font-size:0.72rem;letter-spacing:0.14em;text-transform:uppercase;color:#9ca3af;margin-bottom:0.5rem;'>Executive overview &nbsp;|&nbsp; {today_str}</div>", unsafe_allow_html=True)

    st.markdown(f"""
    <div class="hero-banner" style="border-color:#bbf7d0;background:linear-gradient(135deg,#ecfdf5,#eff6ff);">
        <div class="hero-heading">📈 {t('SWAG Sales Analytics','تحليل مبيعات SWAG')}</div>
        <div class="hero-sub">{t('Understand which models drive revenue and how many days of stock remain.','افهم أي الموديلات تحرك الإيرادات وكم من الأيام تبقى في المخزون.')}</div>
        <div class="hero-tag-row">
            <span class="hero-tag" style="background:#ecfdf3;border-color:#bbf7d0;color:#166534;">Sales Orders</span>
            <span class="hero-tag" style="background:#ecfdf3;border-color:#bbf7d0;color:#166534;">Revenue</span>
            <span class="hero-tag" style="background:#ecfdf3;border-color:#bbf7d0;color:#166534;">Customer Mix</span>
        </div>
    </div>
    """, unsafe_allow_html=True)

    _section_header(t("Filters","الفلاتر"), "⚙️")
    st.markdown("<div class='panel'>", unsafe_allow_html=True)
    col_d1, col_d2, col_mc = st.columns([1, 1, 2])
    with col_d1:
        date_from = st.date_input(t("From Date","من تاريخ"), value=datetime.now()-timedelta(days=30), key="sa_from")
    with col_d2:
        date_to = st.date_input(t("To Date","إلى تاريخ"), value=datetime.now(), key="sa_to")
    with col_mc:
        mc_input = st.text_input(t("Model Code Filter (optional)","فلتر الموديل"), placeholder="e.g. ABC-001", key="sa_mc_filter")
    if st.button(f"🔍 {t('Fetch Sales Data','جلب بيانات المبيعات')}", type="primary", key="sa_fetch", use_container_width=True):
        mc_list = [c.strip().upper() for c in re.split(r"[,\n;]+", mc_input) if c.strip()] if mc_input.strip() else []
        with st.spinner(t("Fetching sales history…","جاري جلب سجل المبيعات…")):
            sa_df = fetch_all_systems_sales_history(mc_list, str(date_from), str(date_to))
        st.session_state.salesanalyticsdf = sa_df
        st.session_state.page_sales = 0
        st.rerun()
    st.markdown("</div>", unsafe_allow_html=True)

    sa_df = st.session_state.get("salesanalyticsdf")
    if sa_df is None or sa_df.empty:
        if st.session_state.salesanalyticsdf is not None:
            st.markdown(f"<div class='empty-state'>ℹ️ {t('No sales data found.','لا توجد بيانات مبيعات.')}</div>", unsafe_allow_html=True)
        return

    _divider()

    # KPIs
    total_rev = _to_num(sa_df[C_SUBTOTAL]).sum() if C_SUBTOTAL in sa_df.columns else 0
    total_qty_s = _to_num(sa_df[C_QTY]).sum() if C_QTY in sa_df.columns else 0
    unique_sos = sa_df[C_SO].nunique() if C_SO in sa_df.columns else 0
    top_system = sa_df.groupby(C_SYSTEM)[C_SUBTOTAL].sum().idxmax() if C_SYSTEM in sa_df.columns and not sa_df.empty else "—"

    _render_kpi_grid([
        _kpi_card("💰", f"SAR {total_rev:,.0f}", t("Total Revenue","إجمالي الإيرادات"), "Selected period", "#22c55e"),
        _kpi_card("📦", f"{total_qty_s:,.0f}", t("Units Sold","الوحدات المباعة"), "All SO lines", "#0ea5e9"),
        _kpi_card("📋", f"{unique_sos:,}", t("Sales Orders","أوامر البيع"), "Confirmed", "#f59e0b"),
        _kpi_card("🏆", str(top_system)[:14], t("Top System","أفضل نظام"), "By revenue", "#a855f7"),
    ])

    # Charts
    col_c1, col_c2 = st.columns(2)
    with col_c1:
        st.markdown("<div class='panel'>", unsafe_allow_html=True)
        st.markdown("<div class='panel-title'>Top Models by Sales</div><div class='panel-sub'>Qty sold per model.</div>", unsafe_allow_html=True)
        fig = _top10_bar(sa_df, C_MODEL, C_SUBTOTAL, "", "#22c55e")
        if fig:
            st.plotly_chart(fig, use_container_width=True)
        st.markdown("</div>", unsafe_allow_html=True)

    with col_c2:
        st.markdown("<div class='panel'>", unsafe_allow_html=True)
        st.markdown("<div class='panel-title'>Sales by Day</div><div class='panel-sub'>Daily revenue line chart.</div>", unsafe_allow_html=True)
        if C_DATE in sa_df.columns:
            daily = sa_df.copy()
            daily["_date"] = pd.to_datetime(daily[C_DATE], errors="coerce").dt.date
            daily_agg = daily.groupby("_date")[C_SUBTOTAL].sum().reset_index()
            daily_agg.columns = ["Date","Revenue"]
            if not daily_agg.empty:
                fig2 = px.area(daily_agg, x="Date", y="Revenue",
                               color_discrete_sequence=["#22c55e"], template="plotly_white")
                fig2.update_traces(fill="tozeroy", fillcolor="rgba(34,197,94,0.10)")
                st.plotly_chart(_apply_light_theme(fig2), use_container_width=True)
        st.markdown("</div>", unsafe_allow_html=True)

    col_c3, col_c4 = st.columns(2)
    with col_c3:
        st.markdown("<div class='panel'>", unsafe_allow_html=True)
        st.markdown("<div class='panel-title'>Customer Mix</div><div class='panel-sub'>Revenue by customer.</div>", unsafe_allow_html=True)
        if C_CUSTOMER in sa_df.columns:
            cust_agg = sa_df.groupby(C_CUSTOMER)[C_SUBTOTAL].sum().reset_index().nlargest(8, C_SUBTOTAL)
            fig_cat = _donut_chart(cust_agg, C_CUSTOMER, C_SUBTOTAL, "",
                                   ["#22c55e","#059669","#0ea5e9","#f59e0b","#ef4444","#a855f7","#f97316","#60a5fa"])
            if fig_cat:
                st.plotly_chart(fig_cat, use_container_width=True)
        st.markdown("</div>", unsafe_allow_html=True)
    with col_c4:
        st.markdown("<div class='panel'>", unsafe_allow_html=True)
        st.markdown("<div class='panel-title'>Revenue by System</div><div class='panel-sub'>System contribution.</div>", unsafe_allow_html=True)
        if C_SYSTEM in sa_df.columns:
            sys_rev = sa_df.groupby(C_SYSTEM)[C_SUBTOTAL].sum().reset_index()
            fig = px.bar(sys_rev, x=C_SYSTEM, y=C_SUBTOTAL,
                         color=C_SUBTOTAL, color_continuous_scale=["#bbf7d0","#22c55e"],
                         template="plotly_white")
            fig.update_traces(marker_cornerradius=5, marker_line_width=0)
            st.plotly_chart(_apply_light_theme(fig), use_container_width=True)
        st.markdown("</div>", unsafe_allow_html=True)

    # Table
    _section_header(t("Sales Detail","تفاصيل المبيعات"), "📋",
                    t("SO data with stock estimates via SWAG backend.","بيانات أوامر البيع."))
    display_df(sa_df, 0, "sa_table")
    sc1, sc2 = st.columns(2)
    with sc1:
        st.download_button(f"⬇️ {t('Export CSV','CSV')}", to_csv(df_for_display(sa_df)), dl_name("sales","csv"), "text/csv", use_container_width=True)
    with sc2:
        st.download_button(f"⬇️ {t('Export Excel','Excel')}", to_excel(df_for_display(sa_df)), dl_name("sales","xlsx"), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

    # Per-code
    _section_header(t("Per-Code Analysis","تحليل لكل رمز"), "🔎")
    codes_in_df = sorted(sa_df[C_MODEL].dropna().unique().tolist()) if C_MODEL in sa_df.columns else []
    if codes_in_df:
        selected_code = st.selectbox(t("Select Model Code","اختر رمز الموديل"), codes_in_df, key="sa_code_select")
        if selected_code:
            code_sales = sa_df[sa_df[C_MODEL] == selected_code]
            total_df_cur = st.session_state.get("total_df")
            code_stock = total_df_cur[total_df_cur[C_MODEL] == selected_code] if total_df_cur is not None else None
            dc1, dc2 = st.columns(2)
            with dc1:
                st.markdown("<div class='panel'>", unsafe_allow_html=True)
                st.markdown(f"<div class='panel-title'>📈 {t('Sales History','سجل المبيعات')}</div>", unsafe_allow_html=True)
                render_premium_table(df_for_display(code_sales))
                st.markdown("</div>", unsafe_allow_html=True)
            with dc2:
                st.markdown("<div class='panel'>", unsafe_allow_html=True)
                st.markdown(f"<div class='panel-title'>📦 {t('Current Stock','المخزون الحالي')}</div>", unsafe_allow_html=True)
                if code_stock is not None and not code_stock.empty:
                    render_premium_table(df_for_display(code_stock[[C_SYSTEM,C_PRODUCT,C_ON_HAND,C_SALE_PRICE]]))
                else:
                    st.markdown(f"<div class='empty-state'>{t('No stock data. Run a stock search first.','لا توجد بيانات مخزون.')}</div>", unsafe_allow_html=True)
                st.markdown("</div>", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 22: MAIN DASHBOARD ROUTER
# ─────────────────────────────────────────────────────────────────────────────
def show_dashboard():
    st.markdown(_build_css(), unsafe_allow_html=True)
    _render_sidebar()
    view = st.session_state.get("analytics_view", "stock")
    if view == "stock":
        show_stock_comparison()
    elif view == "purchase":
        show_purchase_analytics()
    elif view == "sales":
        show_sales_analytics()
    else:
        show_stock_comparison()

# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────
restore_session()
if not st.session_state.authenticated:
    show_login()
else:
    show_dashboard()
